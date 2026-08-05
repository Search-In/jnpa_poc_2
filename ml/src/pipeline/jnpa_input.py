"""
jnpa_input.py -- canonical input adapter for the JNPA UC-1 model suite.
=====================================================================

WHAT THIS FILE IS
-----------------
The eight ``uc1_m*.py`` model files are deliberately self-contained: each one owns
its own dataclasses and will run with nothing but the Python standard library.
That makes them easy to test and to drop into another codebase, but it means
none of them knows how to read *your* spreadsheet.

This module is the missing half. It reads one operational input file -- the
`Vessel_Training_Input_Sample.xlsx` layout, or CSV, or JSON -- validates it,
normalises it, and converts each row into the exact dataclasses M1..M8 expect.

    Vessel_Training_Input_Sample.xlsx
                |
                v
        [ jnpa_input.py ]  <- schema, aliases, validation, IST->UTC, derivations
                |
      +---------+---------+---------+ ... +---------+
      v         v         v                        v
     M1        M2        M3                       M8      (unchanged, still standalone)

Unlike the eight model files, this one is an *integration* layer, so it is
allowed to import them. That is deliberate: duplicating the DUKC core a fifth
time here would add drift risk for no benefit. The models never import this
file, so their standalone property is untouched.

INPUT SCHEMA (the sample workbook's 25 columns)
-----------------------------------------------
    Vessel, IMO, Voyage, ETA, ATA, Import_TEU, Export_TEU, Total_TEU,
    Cargo_Weight_MT, Draft_m, LOA_m, Terminal, Requested_Berth,
    Pilot_Available, Tug_Available, DUKC_Status, Tide_Window_Start,
    Tide_Window_End, Weather, Wind_Speed_kn, Rain_mm_hr, Berth_Occupancy_%,
    Channel_Depth_m, Incident, Cranes_Available

Targets to be produced (never read as inputs): ``ETB``, ``TAT``, ``ETD``.
See ``predict.py``.

Header matching is tolerant: case, spaces, underscores, hyphens and the ``%``
sign are all ignored, so ``Berth Occupancy %``, ``berth_occupancy_pct`` and
``BERTH-OCCUPANCY%`` all resolve to the same field. Unknown columns are
*retained* in ``raw`` and reported as INFO, never silently dropped.

OPTIONAL COLUMNS
----------------
Every derivation this module performs can be overridden by supplying the column
directly. If your data has a real measured value, give it to us and we will use
it instead of the estimate:

    Tide_Height_m, Speed_kn, Vessel_Class, Cargo_Type, Distance_NM,
    Service_Hours, Berth_Ready, Anchorage_Queue, Siltation_m, Dredging_Delta_m,
    Terminal_Max_Draft_m, Bow_Thruster, Priority, Tide_Window_Date

THREE THINGS TO KNOW BEFORE YOU TRUST THE NUMBERS
-------------------------------------------------
1. ``DUKC_Status`` in the sheet is **M1's output, not its input.** It is parsed
   and carried as ``dukc_status_reported`` and used *only* to score the model
   against your data. It is never fed to any model. See ``M1_VALIDATION`` in
   ``run_model.py``.

2. ``Cranes_Available`` is a genuine TAT driver but is **not** in M3's
   ``FEATURE_COLUMNS``. Feeding it would violate the model's import-time
   allow-list assertion. It is carried through to the output as context and
   flagged. Using it properly means adding it to ``FEATURE_COLUMNS`` and
   retraining -- a deliberate, versioned change, not something this adapter
   should do behind your back.

3. The sheet has no tide *height* column, only a tide *window*. Tide height is
   therefore taken from the synthetic harmonic model by default
   (``--tide-policy harmonic``), which is SYNTHETIC, not measured. Supply a
   ``Tide_Height_m`` column and it will be used instead. Every result carries
   ``tide_source`` so this is never invisible.

TIME ZONE
---------
All timestamps in JNPA source documents are IST (UTC+05:30, no DST). They are
converted to timezone-aware UTC here, at the boundary, exactly once. Everything
downstream is UTC. Output timestamps are rendered back to IST for humans, with
the UTC value kept alongside.

CLI
---
    python run.py input --input data/input/Vessel_Training_Input_Sample.xlsx
    python run.py input --input data.xlsx --validate --json
    python run.py input --emit-template my_input.xlsx
    python run.py input --selftest

Exit code is 0 when every check passes and no ERROR-severity issue was found,
1 otherwise -- so this doubles as a CI gate on your data feed.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import sys
import zipfile
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, time, timedelta, timezone
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import jnpa_paths

jnpa_paths.ensure_on_syspath()

# ---------------------------------------------------------------------------
# SECTION 1 -- versioned constants
# ---------------------------------------------------------------------------

MODULE_ID: str = "UC1-IO"
MODULE_VERSION: str = "jnpa-input-v1.0.0"
SCHEMA_VERSION: str = "vessel-call-input/1.0.0"

IST_OFFSET_HOURS: float = 5.5
"""IST is UTC+05:30 and has no daylight saving. Fixed offset is exact."""

IST = timezone(timedelta(hours=IST_OFFSET_HOURS), name="IST")

REF_CHANNEL_DEPTH_M: float = 15.0
"""
Reference charted depth of the controlling reach (CH-INNER), from M1's
DEFAULT_REACHES. ``Channel_Depth_m`` in the sheet is interpreted as the
*surveyed* controlling depth on the day, so::

    net_channel_depth_delta_m = Channel_Depth_m - REF_CHANNEL_DEPTH_M

A negative delta is booked as siltation, a positive delta as dredging gain.
This is the single assumption that links the sheet's depth column to the DUKC
physics; it is stated in every breakdown as ``depth_interpretation``.
"""

PILOT_ROSTER_N: int = 3
TUG_ROSTER_N: int = 4
"""
Roster sizes used to convert the sheet's categorical availability flags into the
integer counts M3/M7/M8 need. These match the WS2 spec's PoC roster (3 pilots /
4 tugs), which is also M8's ``PILOT_AVAIL_N`` / ``TUG_AVAIL_N`` baseline.
Note M7's *real* roster from Details_of_Port_Crafts.pdf is larger (4 pilot
launches, 10 tugs); M7 is given the real roster and these counts are used only
where a 0-3 / 0-4 scale is expected.
"""

# ASSUMED mapping -- the sheet's availability column is categorical, the models
# need a count. Literal reading: "Yes" = full roster, "Busy" = one unit tied up,
# "No" = none available.
PILOT_AVAILABILITY_MAP: Dict[str, int] = {
    "YES": PILOT_ROSTER_N,
    "AVAILABLE": PILOT_ROSTER_N,
    "Y": PILOT_ROSTER_N,
    "BUSY": PILOT_ROSTER_N - 1,
    "PARTIAL": PILOT_ROSTER_N - 1,
    "LIMITED": PILOT_ROSTER_N - 1,
    "NO": 0,
    "N": 0,
    "UNAVAILABLE": 0,
}
TUG_AVAILABILITY_MAP: Dict[str, int] = {
    "YES": TUG_ROSTER_N,
    "AVAILABLE": TUG_ROSTER_N,
    "Y": TUG_ROSTER_N,
    "BUSY": TUG_ROSTER_N - 1,
    "PARTIAL": TUG_ROSTER_N - 1,
    "LIMITED": TUG_ROSTER_N - 1,
    "NO": 0,
    "N": 0,
    "UNAVAILABLE": 0,
}

WEATHER_SEVERITY_MAP: Dict[str, int] = {
    "CLEAR": 0, "FAIR": 0, "FINE": 0, "SUNNY": 0, "CALM": 0,
    "CLOUDY": 1, "OVERCAST": 1, "HAZE": 1, "HAZY": 1, "DRIZZLE": 1,
    "LIGHT RAIN": 1, "LIGHT SHOWERS": 1, "MIST": 1,
    "MODERATE RAIN": 2, "RAIN": 2, "SHOWERS": 2, "ROUGH": 2, "WINDY": 2,
    "SQUALL": 2, "SQUALLY": 2, "THUNDERSTORM": 2,
    "HEAVY RAIN": 3, "STORM": 3, "GALE": 3, "CYCLONE": 3, "MONSOON": 3,
    "VERY ROUGH": 3, "HEAVY SQUALL": 3,
}

INCIDENT_SEVERITY_MAP: Dict[str, int] = {
    "": 0, "NO": 0, "NONE": 0, "NIL": 0, "N": 0, "FALSE": 0,
    "MINOR DELAY": 1, "MINOR": 1, "DELAY": 1, "YES": 1, "Y": 1, "TRUE": 1,
    "MAJOR DELAY": 2, "MAJOR": 2, "BREAKDOWN": 2, "EQUIPMENT FAILURE": 2,
    "STOPPAGE": 3, "SUSPENDED": 3, "ACCIDENT": 3, "CRITICAL": 3,
}

SEVERE_WEATHER_WIND_KN: float = 25.0
SEVERE_WEATHER_RAIN_MMHR: float = 10.0
"""severe_weather_flag fires on severity>=2 OR wind>=25 kn OR rain>=10 mm/hr."""

QUEUE_FROM_OCCUPANCY_BASE_PCT: float = 60.0
QUEUE_FROM_OCCUPANCY_PCT_PER_VESSEL: float = 5.0
"""
DERIVED PROXY, overridable by an ``Anchorage_Queue`` column::

    anchorage_queue = max(0, round((Berth_Occupancy_% - 60) / 5))

Calibrated so the sheet's 68%/72% map to 2 waiting vessels (M8's
``ANCHORAGE_QUEUE_N`` baseline) and 90% maps to 6 (its congested state). This is
an estimate, not a measurement; ``queue_source`` records which was used.
"""

NOMINAL_CALLS_PER_DAY: int = 10
"""JNPA public reference is 10-12 calls/day; extra_arrivals_24h counts above 10."""

VESSEL_CLASS_LOA_BANDS: Tuple[Tuple[float, str], ...] = (
    (350.0, "ULCV"),
    (294.0, "POST_PANAMAX"),
    (225.0, "PANAMAX"),
    (0.0, "FEEDER"),
)

DEFAULT_TRANSIT_SPEED_KN: float = 10.0
"""CH-INNER speed cap. Used when the sheet has no Speed_kn column."""

DEFAULT_DISTANCE_NM: float = 240.0
DEFAULT_SERVICE_HOURS: float = 24.0
DEFAULT_PRIORITY: int = 5

BULK_CARGO_HINTS: Tuple[str, ...] = (
    "BULK", "COAL", "ORE", "GRAIN", "FERTIL", "CEMENT", "CLINKER",
    "LIQUID", "CHEMICAL", "PETROL", "CRUDE", "LPG", "LNG", "OIL",
)

# ---------------------------------------------------------------------------
# SECTION 2 -- column schema and aliases
# ---------------------------------------------------------------------------


def _key(name: Any) -> str:
    """Normalise a header cell to a match key: lowercase alphanumerics only."""
    return re.sub(r"[^a-z0-9]", "", str(name or "").strip().lower())


@dataclass(frozen=True)
class ColumnSpec:
    """One column of the canonical input schema."""

    canonical: str
    field_name: str
    kind: str  # str | int | float | datetime | time | bool
    required: bool
    aliases: Tuple[str, ...]
    unit: str = ""
    description: str = ""
    example: Any = ""

    @property
    def match_keys(self) -> Tuple[str, ...]:
        return tuple({_key(self.canonical), *(_key(a) for a in self.aliases)})


COLUMN_SPECS: Tuple[ColumnSpec, ...] = (
    # --- identity -----------------------------------------------------------
    ColumnSpec("Vessel", "vessel_name", "str", True,
               ("Vessel_Name", "Ship", "Ship_Name", "VESSEL NAME"),
               "", "Vessel name as shown on the call sheet.", "HONG YONG CHANG SHENG"),
    ColumnSpec("IMO", "imo", "str", False,
               ("IMO_No", "IMO_Number", "IMONumber"),
               "", "IMO number; identity only, never a feature.", "1103316"),
    ColumnSpec("Voyage", "voyage", "str", False,
               ("Voyage_No", "Via_No", "VIA", "Voyage_Number", "VCN"),
               "", "Voyage / VIA number.", "10N"),
    # --- timestamps ---------------------------------------------------------
    ColumnSpec("ETA", "eta_ist", "datetime", False,
               ("ETA_IST", "Expected_Arrival", "EDTA"),
               "IST", "Estimated time of arrival (IST).", "2026-07-29 06:00"),
    ColumnSpec("ATA", "ata_ist", "datetime", True,
               ("ATA_IST", "Actual_Arrival", "Arrival"),
               "IST", "Actual time of arrival at anchorage/fairway buoy (IST). "
                      "Anchor for every downstream time calculation.",
               "2026-07-29 07:45"),
    # --- cargo --------------------------------------------------------------
    ColumnSpec("Import_TEU", "import_teu", "int", False, ("Import", "IMP_TEU"),
               "TEU", "Import container parcel.", 2200),
    ColumnSpec("Export_TEU", "export_teu", "int", False, ("Export", "EXP_TEU"),
               "TEU", "Export container parcel.", 1400),
    ColumnSpec("Total_TEU", "total_teu", "int", False, ("TEU", "Parcel_TEU", "Moves"),
               "TEU", "Total parcel. Falls back to Import+Export when blank.", 3600),
    ColumnSpec("Cargo_Weight_MT", "cargo_weight_mt", "float", False,
               ("Cargo_Weight", "Tonnage", "Cargo_MT"),
               "MT", "Cargo weight in metric tonnes.", 52000),
    # --- vessel dimensions --------------------------------------------------
    ColumnSpec("Draft_m", "draft_m", "float", True,
               ("Draft", "Arrival_Draft_m", "Static_Draft_m", "Max_Draft"),
               "m", "Static arrival draft. Drives DUKC and berth compatibility.", 13.2),
    ColumnSpec("LOA_m", "loa_m", "float", False, ("LOA", "Length_Overall", "Length_m"),
               "m", "Length overall; drives berth-length feasibility.", 182.08),
    # --- berth --------------------------------------------------------------
    ColumnSpec("Terminal", "terminal", "str", False,
               ("Terminal_Name", "Operator"),
               "", "Terminal operator code (NSFT/NSICT/NSIGT/APMT/BMCT/...).", "NSICT"),
    ColumnSpec("Requested_Berth", "requested_berth", "str", False,
               ("Berth", "Berth_No", "Berth_Id", "Requested_Berth_Id"),
               "", "Requested berth; canonicalised to the JNPA roster form.", "CB04"),
    # --- resources ----------------------------------------------------------
    ColumnSpec("Pilot_Available", "pilot_available_raw", "str", False,
               ("Pilot", "Pilots_Available", "Pilot_Avail"),
               "", "Yes / Busy / No. Mapped to a pilot count.", "Yes"),
    ColumnSpec("Tug_Available", "tug_available_raw", "str", False,
               ("Tug", "Tugs_Available", "Tug_Avail"),
               "", "Yes / Busy / No. Mapped to a tug count.", "Yes"),
    ColumnSpec("Cranes_Available", "cranes_available", "int", False,
               ("Cranes", "QC_Available", "Quay_Cranes"),
               "count", "Quay cranes allocated. CONTEXT ONLY -- not an M3 feature; "
                        "see the module docstring.", 4),
    # --- observed DUKC status (LABEL, not input) ----------------------------
    ColumnSpec("DUKC_Status", "dukc_status_reported", "str", False,
               ("UKC_Status", "DUKC", "Under_Keel_Status"),
               "", "OBSERVED status from your records. Used to SCORE M1, never "
                   "fed into it.", "Safe"),
    # --- tide ---------------------------------------------------------------
    ColumnSpec("Tide_Window_Start", "tide_window_start_raw", "time", False,
               ("Window_Start", "Tidal_Window_Start"),
               "IST HH:MM", "Start of the stated tidal window (time of day, IST).", "06:30"),
    ColumnSpec("Tide_Window_End", "tide_window_end_raw", "time", False,
               ("Window_End", "Tidal_Window_End"),
               "IST HH:MM", "End of the stated tidal window. If earlier than the "
                            "start it is taken to cross midnight.", "09:30"),
    # --- weather ------------------------------------------------------------
    ColumnSpec("Weather", "weather_raw", "str", False, ("Weather_Condition", "Wx"),
               "", "Free text; mapped to a 0-3 severity.", "Clear"),
    ColumnSpec("Wind_Speed_kn", "wind_kn", "float", False,
               ("Wind", "Wind_Speed", "Wind_kt", "Wind_Knots"),
               "kn", "Sustained wind speed.", 8),
    ColumnSpec("Rain_mm_hr", "rain_mm_hr", "float", False,
               ("Rain", "Rainfall", "Rain_mm", "Precip_mm_hr"),
               "mm/hr", "Rainfall intensity.", 0),
    # --- port state ---------------------------------------------------------
    ColumnSpec("Berth_Occupancy_%", "berth_occupancy_pct", "float", False,
               ("Berth_Occupancy", "Occupancy", "Occupancy_Pct", "Berth_Utilisation"),
               "%", "Port-wide berth occupancy at the time of the call.", 68),
    ColumnSpec("Channel_Depth_m", "channel_depth_m", "float", False,
               ("Channel_Depth", "Controlling_Depth_m", "Surveyed_Depth_m", "Depth"),
               "m", "Surveyed controlling depth. Delta vs 15.0 m becomes "
                    "siltation (negative) or dredging gain (positive).", 15.4),
    ColumnSpec("Incident", "incident_raw", "str", False,
               ("Incident_Type", "Disruption", "Event"),
               "", "Free text; mapped to a 0-3 severity.", "No"),
    # ----------------------------------------------------------------------
    # OPTIONAL columns -- supply any of these and the derivation is skipped.
    # ----------------------------------------------------------------------
    ColumnSpec("Tide_Height_m", "tide_height_m", "float", False,
               ("Tide", "Tide_m", "Tide_Height", "Water_Level_m"),
               "m", "MEASURED tide height above chart datum at ATA. Overrides the "
                    "synthetic harmonic model.", ""),
    ColumnSpec("Speed_kn", "transit_speed_kn", "float", False,
               ("Transit_Speed_kn", "Speed", "Channel_Speed_kn"),
               "kn", "Planned channel transit speed. Defaults to the reach cap.", ""),
    ColumnSpec("Vessel_Class", "vessel_class_raw", "str", False,
               ("Class", "Ship_Class", "Size_Class"),
               "", "ULCV / POST_PANAMAX / PANAMAX / FEEDER. Derived from LOA "
                   "when blank.", ""),
    ColumnSpec("Cargo_Type", "cargo_type", "str", False,
               ("Cargo", "Commodity", "Cargo_Commodity"),
               "", "Container / Bulk. Selects the block coefficient Cb.", ""),
    ColumnSpec("Distance_NM", "distance_nm", "float", False,
               ("Distance", "DTG", "Distance_To_Go_NM"),
               "NM", "Distance to go at the JIT decision point (M6).", ""),
    ColumnSpec("Service_Hours", "service_hours", "float", False,
               ("Service_H", "Berth_Hours", "Planned_Berth_Hours"),
               "h", "Planned time alongside (M5). Defaults to 24 h.", ""),
    ColumnSpec("Berth_Ready", "berth_ready_ist", "datetime", False,
               ("Berth_Ready_Time", "Berth_Available_From", "ETB_Planned"),
               "IST", "Berth-ready time for the JIT RTA calculation (M6).", ""),
    ColumnSpec("Anchorage_Queue", "anchorage_queue_count", "int", False,
               ("Queue", "Waiting_Vessels", "Anchorage_Count"),
               "vessels", "MEASURED anchorage queue. Overrides the occupancy proxy.", ""),
    ColumnSpec("Siltation_m", "siltation_m", "float", False,
               ("Siltation", "Silt_m"),
               "m", "Depth lost to siltation. Overrides the Channel_Depth delta.", ""),
    ColumnSpec("Dredging_Delta_m", "dredging_delta_m", "float", False,
               ("Dredging", "Dredge_m", "Dredging_Gain_m"),
               "m", "Depth gained by dredging. Overrides the Channel_Depth delta.", ""),
    ColumnSpec("Terminal_Max_Draft_m", "terminal_max_draft_m", "float", False,
               ("Berth_Max_Draft_m", "Max_Draft_m"),
               "m", "Berth declared max draft. Looked up from the roster when blank.", ""),
    ColumnSpec("Bow_Thruster", "bow_thruster_raw", "str", False,
               ("Bow_Thruster_YN", "BT"),
               "", "Y/N. Drives the extra-tug rule in M7.", ""),
    ColumnSpec("Priority", "priority", "int", False, ("Prio", "Rank"),
               "1-9", "Berthing priority for M5. 1 = highest. Defaults to 5.", ""),
    ColumnSpec("Tide_Window_Date", "tide_window_date_raw", "datetime", False,
               ("Window_Date",),
               "IST", "Date the tide window belongs to. Defaults to the ATA date.", ""),
)

SPEC_BY_FIELD: Dict[str, ColumnSpec] = {s.field_name: s for s in COLUMN_SPECS}

_MATCH_INDEX: Dict[str, ColumnSpec] = {}
for _s in COLUMN_SPECS:
    for _k in _s.match_keys:
        # First spec wins; later collisions are a schema bug and are asserted below.
        assert _k not in _MATCH_INDEX or _MATCH_INDEX[_k] is _s, (
            f"alias collision on {_k!r} between "
            f"{_MATCH_INDEX[_k].canonical} and {_s.canonical}"
        )
        _MATCH_INDEX[_k] = _s

REQUIRED_COLUMNS: Tuple[str, ...] = tuple(s.canonical for s in COLUMN_SPECS if s.required)

TARGET_COLUMNS: Tuple[str, ...] = ("ETB", "TAT", "ETD")
"""
Produced by the models, never read as inputs. If these appear in the input file
they are reported as an ERROR -- silently ignoring a target column is how leakage
gets into a pipeline.
"""
_TARGET_KEYS = {_key(c) for c in TARGET_COLUMNS} | {
    _key(x) for x in ("ETB_IST", "ETD_IST", "TAT_Hours", "TAT_h", "Actual_TAT",
                      "ATB", "ATD", "Actual_Berthing", "Actual_Departure")
}

# ---------------------------------------------------------------------------
# SECTION 3 -- value parsers
# ---------------------------------------------------------------------------

_DATETIME_FORMATS: Tuple[str, ...] = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
    "%Y-%m-%d",
    "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%d.%m.%Y %H:%M", "%d.%m.%Y",
    "%d-%b-%Y %H:%M", "%d-%b-%Y", "%d %b %Y %H:%M", "%d %b %Y",
    "%m/%d/%Y %H:%M", "%m/%d/%Y",
    "%d%m%Y:%H:%M",  # BERMAN XML literal form, e.g. 11022026:17:00
)

_TIME_FORMATS: Tuple[str, ...] = ("%H:%M:%S", "%H:%M", "%H%M", "%I:%M %p", "%I:%M%p")


class ParseError(ValueError):
    """Raised by the parsers; carries a machine-readable ``code``."""

    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return str(v).strip() in ("", "-", "--", "NA", "N/A", "na", "n/a", "None", "null")


def parse_float(v: Any, *, lo: Optional[float] = None, hi: Optional[float] = None) -> Optional[float]:
    if _is_blank(v):
        return None
    if isinstance(v, bool):
        raise ParseError("not_a_number", f"boolean {v!r} where a number was expected")
    if isinstance(v, (int, float)):
        out = float(v)
    else:
        text = str(v).strip().replace(",", "").replace("%", "")
        # Strip a trailing unit, e.g. "13.2 m" / "8 kn".
        text = re.sub(r"\s*(m|km|kn|kt|kts|knots|mt|teu|hrs?|h|mm/hr|mm)$", "", text, flags=re.I)
        try:
            out = float(text)
        except ValueError as exc:
            raise ParseError("not_a_number", f"cannot read {v!r} as a number") from exc
    if not math.isfinite(out):
        raise ParseError("not_finite", f"{v!r} is not finite")
    if lo is not None and out < lo:
        raise ParseError("below_range", f"{out} is below the allowed minimum {lo}")
    if hi is not None and out > hi:
        raise ParseError("above_range", f"{out} is above the allowed maximum {hi}")
    return out


def parse_int(v: Any, *, lo: Optional[int] = None, hi: Optional[int] = None) -> Optional[int]:
    f = parse_float(v, lo=lo, hi=hi)
    if f is None:
        return None
    if abs(f - round(f)) > 1e-9:
        raise ParseError("not_an_integer", f"{v!r} is not a whole number")
    return int(round(f))


def parse_datetime_ist(v: Any) -> Optional[datetime]:
    """Parse an IST timestamp from a cell and return it as timezone-aware UTC."""
    if _is_blank(v):
        return None
    if isinstance(v, datetime):
        naive = v.replace(tzinfo=None) if v.tzinfo is not None else v
        if v.tzinfo is not None:
            return v.astimezone(timezone.utc)
        return naive.replace(tzinfo=IST).astimezone(timezone.utc)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, tzinfo=IST).astimezone(timezone.utc)
    text = str(v).strip()
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=IST).astimezone(timezone.utc)
        except ValueError:
            continue
    try:  # last resort: ISO-8601 with an explicit offset
        parsed = datetime.fromisoformat(text)
    except ValueError as exc:
        raise ParseError("bad_datetime", f"cannot read {v!r} as a date/time") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=IST)
    return parsed.astimezone(timezone.utc)


def parse_time_of_day(v: Any) -> Optional[time]:
    """Parse an ``HH:MM`` cell. Excel may hand us a ``time`` or a ``datetime``."""
    if _is_blank(v):
        return None
    if isinstance(v, time):
        return v.replace(tzinfo=None)
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, (int, float)) and 0.0 <= float(v) < 1.0:
        # Excel stores a bare time as a fraction of a day.
        total = int(round(float(v) * 86400))
        return time(total // 3600 % 24, total % 3600 // 60, total % 60)
    text = str(v).strip()
    for fmt in _TIME_FORMATS:
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ParseError("bad_time", f"cannot read {v!r} as a time of day")


def parse_choice(v: Any, mapping: Mapping[str, int], *, label: str) -> Optional[int]:
    if _is_blank(v):
        return None
    text = re.sub(r"\s+", " ", str(v).strip()).upper()
    if text in mapping:
        return mapping[text]
    # Substring fallback, longest key first so "HEAVY RAIN" beats "RAIN".
    for keyword in sorted(mapping, key=len, reverse=True):
        if keyword and keyword in text:
            return mapping[keyword]
    raise ParseError("unknown_category", f"unrecognised {label} value {v!r}")


def parse_bool(v: Any, default: Optional[bool] = None) -> Optional[bool]:
    if _is_blank(v):
        return default
    text = str(v).strip().upper()
    if text in ("Y", "YES", "TRUE", "1", "FITTED", "AVAILABLE"):
        return True
    if text in ("N", "NO", "FALSE", "0", "NOT FITTED", "NIL"):
        return False
    raise ParseError("bad_boolean", f"cannot read {v!r} as yes/no")


# --- berth / terminal canonicalisation --------------------------------------

TERMINAL_ALIASES: Dict[str, str] = {
    "JNPCT": "NSFT",       # renamed when the terminal was privatised
    "JNPT": "NSFT",
    "NSFT": "NSFT",
    "NSICT": "NSICT",
    "DPWORLDNSICT": "NSICT",
    "NSIGT": "NSIGT",
    "DPWORLDNSIGT": "NSIGT",
    "GTI": "APMT",
    "APMT": "APMT",
    "APMTERMINALS": "APMT",
    "BMCT": "BMCT",
    "PSABMCT": "BMCT",
    "NSDT": "NSDT",
    "BPCL": "BPCL",
    "JJLTPL": "JJLTPL",
}

BERTH_PREFIX_BY_TERMINAL: Dict[str, str] = {
    "NSFT": "CB", "NSICT": "CB", "NSIGT": "CB",
    "APMT": "APMT", "BMCT": "BMCT", "NSDT": "NSD",
    "BPCL": "LB", "JJLTPL": "LB",
}

BERTH_SHORTCODE_ALIASES: Dict[str, str] = {
    # Short forms seen in operational sheets -> canonical roster id.
    "BM": "BMCT", "BMC": "BMCT", "BMCT": "BMCT",
    "AP": "APMT", "APM": "APMT", "APMT": "APMT",
    "CB": "CB", "CCB": "CCB", "LB": "LB", "NSD": "NSD",
}


def canonical_terminal(raw: Any) -> str:
    key = re.sub(r"[^A-Z0-9]", "", str(raw or "").upper())
    return TERMINAL_ALIASES.get(key, str(raw or "").strip().upper())


def canonical_berth(raw: Any, terminal: str = "") -> str:
    """
    ``"CB04"`` -> ``"CB-04"``; ``"BM05"`` -> ``"BMCT-05"``; ``"CCB-N"`` -> unchanged.

    The terminal is used only to disambiguate a bare number (``"5"`` at BMCT
    becomes ``BMCT-05``). Unknown ids are upper-cased and returned as-is so a
    new berth never silently vanishes.
    """
    text = re.sub(r"\s+", "", str(raw or "").upper())
    if not text:
        return ""
    if text in ("CCB-N", "CCBN", "CCB-S", "CCBS"):
        return "CCB-" + text[-1]
    m = re.match(r"^([A-Z]+)[-_ ]?(\d+)$", text)
    if m:
        prefix, number = m.group(1), int(m.group(2))
        prefix = BERTH_SHORTCODE_ALIASES.get(prefix, prefix)
        return f"{prefix}-{number:02d}"
    m = re.match(r"^(\d+)$", text)
    if m and terminal:
        prefix = BERTH_PREFIX_BY_TERMINAL.get(canonical_terminal(terminal), "")
        if prefix:
            return f"{prefix}-{int(m.group(1)):02d}"
    return text


def derive_vessel_class(loa_m: Optional[float], total_teu: Optional[int]) -> str:
    if total_teu is not None and total_teu >= 14000:
        return "ULCV"
    if loa_m is None:
        return "POST_PANAMAX"
    for threshold, label in VESSEL_CLASS_LOA_BANDS:
        if loa_m >= threshold:
            return label
    return "FEEDER"


def derive_cargo_class(cargo_type: str) -> str:
    """Return ``CONTAINER`` or ``BULK`` -- this selects the block coefficient."""
    text = str(cargo_type or "").upper()
    if any(hint in text for hint in BULK_CARGO_HINTS):
        return "BULK"
    return "CONTAINER"


# ---------------------------------------------------------------------------
# SECTION 4 -- issues, the canonical row, and the batch
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class Issue:
    """One validation finding. ``ERROR`` blocks the row; ``WARN``/``INFO`` do not."""

    row: int
    column: str
    severity: str  # ERROR | WARN | INFO
    code: str
    message: str
    value: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return asdict(self)

    def __str__(self) -> str:
        where = f"row {self.row}" if self.row > 0 else "file"
        col = f" [{self.column}]" if self.column else ""
        val = f" (got {self.value!r})" if self.value else ""
        return f"{self.severity:<5} {where}{col}: {self.message}{val}"


@dataclass
class VesselCallInput:
    """
    One validated, normalised vessel call.

    Every timestamp is timezone-aware UTC. Every ``*_source`` field records
    whether a value was measured (from your file) or derived (by this adapter),
    so nothing estimated can be mistaken for something observed.
    """

    # --- provenance ---------------------------------------------------------
    row: int
    call_id: str
    source_file: str = ""

    # --- identity -----------------------------------------------------------
    vessel_name: str = ""
    vessel_id: str = ""
    imo: str = ""
    voyage: str = ""

    # --- timestamps (UTC) ---------------------------------------------------
    eta_utc: Optional[datetime] = None
    ata_utc: Optional[datetime] = None
    berth_ready_utc: Optional[datetime] = None
    tide_window_start_utc: Optional[datetime] = None
    tide_window_end_utc: Optional[datetime] = None

    # --- cargo --------------------------------------------------------------
    import_teu: int = 0
    export_teu: int = 0
    total_teu: int = 0
    cargo_weight_mt: float = 0.0
    cargo_type: str = ""
    cargo_class: str = "CONTAINER"

    # --- vessel -------------------------------------------------------------
    draft_m: float = 0.0
    loa_m: float = 0.0
    vessel_class: str = "POST_PANAMAX"
    transit_speed_kn: float = DEFAULT_TRANSIT_SPEED_KN
    has_bow_thruster: bool = True

    # --- berth --------------------------------------------------------------
    terminal: str = ""
    requested_berth: str = ""
    terminal_max_draft_m: float = 0.0
    service_hours: float = DEFAULT_SERVICE_HOURS
    priority: int = DEFAULT_PRIORITY

    # --- resources ----------------------------------------------------------
    pilots_available: int = PILOT_ROSTER_N
    pilots_down: int = 0
    tugs_available: int = TUG_ROSTER_N
    tugs_down: int = 0
    cranes_available: int = 0

    # --- environment --------------------------------------------------------
    weather_raw: str = ""
    weather_severity: int = 0
    severe_weather_flag: int = 0
    wind_kn: float = 0.0
    rain_mm_hr: float = 0.0
    tide_height_m: float = 0.0

    # --- channel ------------------------------------------------------------
    channel_depth_m: float = REF_CHANNEL_DEPTH_M
    net_channel_depth_delta_m: float = 0.0
    siltation_m: float = 0.0
    dredging_delta_m: float = 0.0

    # --- port state ---------------------------------------------------------
    berth_occupancy_pct: float = 0.0
    anchorage_queue_count: int = 0
    incident_raw: str = ""
    incident_severity: int = 0

    # --- batch-derived ------------------------------------------------------
    calls_prev_24h: int = 0
    extra_arrivals_24h: int = 0
    berth_window_extension_h: float = 0.0

    # --- M6 -----------------------------------------------------------------
    distance_nm: float = DEFAULT_DISTANCE_NM

    # --- observed label (NEVER an input) ------------------------------------
    dukc_status_reported: str = ""

    # --- provenance flags ---------------------------------------------------
    tide_source: str = "SYNTHETIC_HARMONIC_v1"
    queue_source: str = "DERIVED_FROM_OCCUPANCY"
    depth_source: str = "COLUMN_Channel_Depth_m"
    speed_source: str = "DEFAULT_REACH_CAP"
    distance_source: str = "DEFAULT"

    raw: Dict[str, Any] = field(default_factory=dict)
    issues: List[Issue] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not any(i.severity == "ERROR" for i in self.issues)

    def as_dict(self) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for k, v in asdict(self).items():
            if k in ("raw", "issues"):
                continue
            out[k] = v.isoformat() if isinstance(v, datetime) else v
        out["issues"] = [i.as_dict() for i in self.issues]
        return out


@dataclass
class InputBatch:
    """A validated file: the rows, the issues, and how it was read."""

    schema_version: str
    source_file: str
    source_format: str
    sheet_name: str
    header_map: Dict[str, str]
    unknown_columns: Tuple[str, ...]
    missing_required: Tuple[str, ...]
    rows: List[VesselCallInput]
    issues: List[Issue]
    read_at_utc: datetime

    @property
    def valid_rows(self) -> List[VesselCallInput]:
        return [r for r in self.rows if r.ok]

    @property
    def error_count(self) -> int:
        return sum(1 for i in self.all_issues if i.severity == "ERROR")

    @property
    def warn_count(self) -> int:
        return sum(1 for i in self.all_issues if i.severity == "WARN")

    @property
    def all_issues(self) -> List[Issue]:
        out = list(self.issues)
        for r in self.rows:
            out.extend(r.issues)
        return out

    @property
    def ok(self) -> bool:
        return self.error_count == 0 and len(self.valid_rows) > 0

    def summary(self) -> Dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "source_file": self.source_file,
            "source_format": self.source_format,
            "sheet_name": self.sheet_name,
            "rows_read": len(self.rows),
            "rows_valid": len(self.valid_rows),
            "errors": self.error_count,
            "warnings": self.warn_count,
            "columns_matched": len(self.header_map),
            "unknown_columns": list(self.unknown_columns),
            "missing_required": list(self.missing_required),
            "read_at_utc": self.read_at_utc.isoformat(),
        }


# ---------------------------------------------------------------------------
# SECTION 5 -- file readers (.xlsx / .csv / .json)
# ---------------------------------------------------------------------------

try:  # openpyxl is the preferred xlsx reader when present
    import openpyxl  # type: ignore

    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover - openpyxl is optional
    openpyxl = None  # type: ignore
    _HAS_OPENPYXL = False


_XLSX_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XLSX_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

_BUILTIN_DATE_FORMATS = set(range(14, 23)) | {27, 30, 36, 45, 46, 47, 50, 57}


def _excel_serial_to_datetime(serial: float) -> datetime:
    """
    Convert an Excel serial date to a naive datetime.

    Excel's 1900 system wrongly treats 1900 as a leap year, so serials above 59
    are one day ahead of the true count. Using 1899-12-30 as the epoch absorbs
    that off-by-one for every date after 1900-03-01, which is every date we will
    ever see in a port dataset.
    """
    return datetime(1899, 12, 30) + timedelta(days=float(serial))


def _read_xlsx_stdlib(path: str, sheet: Optional[str] = None) -> Tuple[str, List[List[Any]]]:
    """
    Minimal .xlsx reader built on ``zipfile`` + ``xml.etree`` only.

    Exists so this adapter keeps the suite's zero-dependency property: if
    openpyxl is not installed, reading the workbook still works. It handles
    shared strings, inline strings, booleans and date-formatted numerics, which
    covers the JNPA input layout.
    """
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(path) as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = [
            (s.get("name") or "", s.get(f"{_XLSX_REL_NS}id") or "")
            for s in wb.iter(f"{_XLSX_NS}sheet")
        ]
        if not sheets:
            raise ValueError(f"{path}: workbook contains no sheets")

        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_target = {
            r.get("Id"): (r.get("Target") or "")
            for r in rels_root
        }

        chosen = None
        for name, rid in sheets:
            if sheet is None or _key(name) == _key(sheet):
                chosen = (name, rid)
                break
        if chosen is None:
            raise ValueError(
                f"{path}: sheet {sheet!r} not found; available: "
                + ", ".join(n for n, _ in sheets)
            )
        sheet_name, rid = chosen
        target = rel_target.get(rid, "worksheets/sheet1.xml").lstrip("/")
        member = target if target.startswith("xl/") else f"xl/{target}"

        shared: List[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            sst = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in sst.iter(f"{_XLSX_NS}si"):
                shared.append("".join(t.text or "" for t in si.iter(f"{_XLSX_NS}t")))

        date_styles: set = set()
        if "xl/styles.xml" in zf.namelist():
            st = ET.fromstring(zf.read("xl/styles.xml"))
            custom_date_ids = {
                int(nf.get("numFmtId") or -1)
                for nf in st.iter(f"{_XLSX_NS}numFmt")
                if re.search(r"[yYdDhHsS]", (nf.get("formatCode") or "").split(";")[0])
                and "General" not in (nf.get("formatCode") or "")
            }
            xfs = st.find(f"{_XLSX_NS}cellXfs")
            if xfs is not None:
                for idx, xf in enumerate(xfs.iter(f"{_XLSX_NS}xf")):
                    fmt_id = int(xf.get("numFmtId") or 0)
                    if fmt_id in _BUILTIN_DATE_FORMATS or fmt_id in custom_date_ids:
                        date_styles.add(idx)

        ws = ET.fromstring(zf.read(member))

    def col_index(ref: str) -> int:
        letters = re.match(r"^([A-Z]+)", ref or "A")
        if not letters:
            return 0
        n = 0
        for ch in letters.group(1):
            n = n * 26 + (ord(ch) - 64)
        return n - 1

    grid: List[List[Any]] = []
    for row_el in ws.iter(f"{_XLSX_NS}row"):
        cells: Dict[int, Any] = {}
        for c in row_el.iter(f"{_XLSX_NS}c"):
            ref, ctype, style = c.get("r") or "", c.get("t") or "n", c.get("s")
            v_el = c.find(f"{_XLSX_NS}v")
            if ctype == "inlineStr":
                is_el = c.find(f"{_XLSX_NS}is")
                text = "".join(t.text or "" for t in is_el.iter(f"{_XLSX_NS}t")) if is_el is not None else ""
                cells[col_index(ref)] = text
                continue
            if v_el is None or v_el.text is None:
                continue
            raw = v_el.text
            if ctype == "s":
                idx = int(raw)
                value: Any = shared[idx] if 0 <= idx < len(shared) else ""
            elif ctype == "b":
                value = raw == "1"
            elif ctype in ("str", "e"):
                value = raw
            else:
                try:
                    num = float(raw)
                except ValueError:
                    value = raw
                else:
                    if style is not None and int(style) in date_styles:
                        value = _excel_serial_to_datetime(num)
                    else:
                        value = int(num) if num == int(num) and abs(num) < 1e15 else num
            cells[col_index(ref)] = value
        if cells:
            width = max(cells) + 1
            grid.append([cells.get(i) for i in range(width)])
        else:
            grid.append([])
    return sheet_name, grid


def read_table(path: str, sheet: Optional[str] = None) -> Tuple[str, str, List[List[Any]]]:
    """
    Read any supported input file into ``(format, sheet_name, grid)``.

    ``grid[0]`` is the header row. Supported: ``.xlsx``/``.xlsm``, ``.csv``,
    ``.json`` (a list of objects, or ``{"rows": [...]}``).
    """
    ext = os.path.splitext(path)[1].lower()
    if not os.path.exists(path):
        raise FileNotFoundError(f"input file not found: {path}")

    if ext in (".xlsx", ".xlsm"):
        if _HAS_OPENPYXL:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            try:
                ws = wb[sheet] if sheet else wb.worksheets[0]
                grid = [list(r) for r in ws.iter_rows(values_only=True)]
                return "xlsx(openpyxl)", ws.title, grid
            finally:
                wb.close()
        name, grid = _read_xlsx_stdlib(path, sheet)
        return "xlsx(stdlib)", name, grid

    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            return "csv", os.path.basename(path), [list(r) for r in csv.reader(fh)]

    if ext == ".json":
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        records = payload.get("rows", payload) if isinstance(payload, dict) else payload
        if not isinstance(records, list) or not records:
            raise ValueError(f"{path}: expected a non-empty list of row objects")
        headers: List[str] = []
        for rec in records:
            for k in rec:
                if k not in headers:
                    headers.append(k)
        grid = [headers] + [[rec.get(h) for h in headers] for rec in records]
        return "json", os.path.basename(path), grid

    raise ValueError(f"unsupported input format {ext!r}; use .xlsx, .csv or .json")


def _find_header_row(grid: Sequence[Sequence[Any]]) -> int:
    """
    Locate the header row.

    Some operational exports carry a title line or a blank line above the real
    header, so we score the first few rows by how many canonical columns they
    match rather than assuming row 0.
    """
    best_idx, best_score = 0, -1
    for idx, row in enumerate(grid[:10]):
        score = sum(1 for cell in row if _key(cell) in _MATCH_INDEX)
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx if best_score > 0 else 0


# ---------------------------------------------------------------------------
# SECTION 6 -- tide provider
# ---------------------------------------------------------------------------
#
# Reuses M2's provider rather than duplicating the harmonic a fifth time. This
# file is an integration layer, so importing a model is correct here; the models
# themselves still import nothing from each other.

def _tide_provider():
    import uc1_m2_tidal_window as m2

    return m2.SyntheticTideProvider()


def harmonic_tide_m(when: datetime) -> float:
    """Synthetic tide height (m above chart datum). SYNTHETIC -- not measured."""
    return _tide_provider().height_m(when)


# ---------------------------------------------------------------------------
# SECTION 7 -- row parsing and derivation
# ---------------------------------------------------------------------------

_RANGES: Dict[str, Tuple[Optional[float], Optional[float]]] = {
    "draft_m": (0.5, 25.0),
    "loa_m": (10.0, 500.0),
    "wind_kn": (0.0, 150.0),
    "rain_mm_hr": (0.0, 300.0),
    "berth_occupancy_pct": (0.0, 100.0),
    "channel_depth_m": (5.0, 30.0),
    "tide_height_m": (-2.0, 8.0),
    "transit_speed_kn": (0.5, 25.0),
    "cargo_weight_mt": (0.0, 500000.0),
    "distance_nm": (0.1, 12000.0),
    "service_hours": (0.5, 240.0),
    "terminal_max_draft_m": (5.0, 25.0),
    "import_teu": (0, 40000),
    "export_teu": (0, 40000),
    "total_teu": (0, 60000),
    "cranes_available": (0, 20),
    "anchorage_queue_count": (0, 200),
    "priority": (1, 9),
    "siltation_m": (0.0, 5.0),
    "dredging_delta_m": (0.0, 5.0),
}


def _terminal_max_draft_lookup(berth_id: str, terminal: str) -> Optional[float]:
    try:
        import uc1_m4_berth_utilisation as m4
    except Exception:  # pragma: no cover
        return None
    by_id = {b.berth_id: b for b in m4.default_berths()}
    if berth_id in by_id:
        return float(by_id[berth_id].max_draft_m)
    same_terminal = [b for b in m4.default_berths() if b.terminal == terminal]
    if same_terminal:
        return float(max(b.max_draft_m for b in same_terminal))
    return None


def _parse_row(
    row_no: int,
    values: Mapping[str, Any],
    *,
    tide_policy: str,
    fixed_tide_m: Optional[float],
    source_file: str,
) -> VesselCallInput:
    """Turn one raw row into a validated ``VesselCallInput``."""
    issues: List[Issue] = []
    got: Dict[str, Any] = {}

    def take(field_name: str, kind: Optional[str] = None) -> Any:
        spec = SPEC_BY_FIELD[field_name]
        raw_value = values.get(field_name)
        if _is_blank(raw_value):
            if spec.required:
                issues.append(Issue(row_no, spec.canonical, "ERROR", "missing_required",
                                    f"{spec.canonical} is required but blank"))
            return None
        lo, hi = _RANGES.get(field_name, (None, None))
        try:
            k = kind or spec.kind
            if k == "float":
                return parse_float(raw_value, lo=lo, hi=hi)
            if k == "int":
                return parse_int(raw_value,
                                 lo=None if lo is None else int(lo),
                                 hi=None if hi is None else int(hi))
            if k == "datetime":
                return parse_datetime_ist(raw_value)
            if k == "time":
                return parse_time_of_day(raw_value)
            return str(raw_value).strip()
        except ParseError as exc:
            severity = "ERROR" if spec.required else "WARN"
            issues.append(Issue(row_no, spec.canonical, severity, exc.code,
                                str(exc), str(raw_value)))
            return None

    for spec in COLUMN_SPECS:
        got[spec.field_name] = take(spec.field_name)

    # --- identity -----------------------------------------------------------
    vessel_name = str(got.get("vessel_name") or f"VESSEL-{row_no}")
    imo = str(got.get("imo") or "").strip()
    voyage = str(got.get("voyage") or "").strip()
    vessel_id = imo or re.sub(r"[^A-Z0-9]", "", vessel_name.upper())[:12] or f"V{row_no:04d}"
    call_id = f"C-{row_no:04d}"

    # --- cargo --------------------------------------------------------------
    import_teu = int(got.get("import_teu") or 0)
    export_teu = int(got.get("export_teu") or 0)
    total_teu = got.get("total_teu")
    if total_teu is None:
        total_teu = import_teu + export_teu
        if total_teu > 0:
            issues.append(Issue(row_no, "Total_TEU", "INFO", "derived",
                                "Total_TEU blank; used Import_TEU + Export_TEU"))
    total_teu = int(total_teu)
    if import_teu and export_teu and total_teu and abs((import_teu + export_teu) - total_teu) > 1:
        issues.append(Issue(row_no, "Total_TEU", "WARN", "teu_mismatch",
                            f"Import({import_teu}) + Export({export_teu}) "
                            f"= {import_teu + export_teu} but Total_TEU = {total_teu}"))

    cargo_type = str(got.get("cargo_type") or "")
    cargo_class = derive_cargo_class(cargo_type)

    # --- vessel -------------------------------------------------------------
    draft_m = float(got.get("draft_m") or 0.0)
    loa_m = float(got.get("loa_m") or 0.0)
    vessel_class = str(got.get("vessel_class_raw") or "").strip().upper().replace("-", "_")
    if vessel_class not in ("ULCV", "POST_PANAMAX", "PANAMAX", "FEEDER"):
        if vessel_class:
            issues.append(Issue(row_no, "Vessel_Class", "WARN", "unknown_class",
                                f"unrecognised class {vessel_class!r}; derived from LOA"))
        vessel_class = derive_vessel_class(loa_m or None, total_teu or None)

    speed = got.get("transit_speed_kn")
    speed_source = "COLUMN_Speed_kn" if speed is not None else "DEFAULT_REACH_CAP"
    transit_speed_kn = float(speed if speed is not None else DEFAULT_TRANSIT_SPEED_KN)

    try:
        bow_thruster = parse_bool(got.get("bow_thruster_raw"), default=True)
    except ParseError as exc:
        issues.append(Issue(row_no, "Bow_Thruster", "WARN", exc.code, str(exc)))
        bow_thruster = True

    # --- timestamps ---------------------------------------------------------
    eta_utc = got.get("eta_ist")
    ata_utc = got.get("ata_ist")
    if eta_utc and ata_utc and ata_utc < eta_utc - timedelta(hours=72):
        issues.append(Issue(row_no, "ATA", "WARN", "ata_far_before_eta",
                            "ATA is more than 72 h before ETA -- check the date fields"))

    # --- tide window: HH:MM anchored to a date ------------------------------
    anchor_utc = got.get("tide_window_date_raw") or ata_utc or eta_utc
    tw_start_utc = tw_end_utc = None
    if anchor_utc is not None:
        anchor_ist_date = anchor_utc.astimezone(IST).date()
        ws, we = got.get("tide_window_start_raw"), got.get("tide_window_end_raw")
        if ws is not None:
            tw_start_utc = datetime.combine(anchor_ist_date, ws, tzinfo=IST).astimezone(timezone.utc)
        if we is not None:
            base = tw_start_utc.astimezone(IST).date() if tw_start_utc else anchor_ist_date
            tw_end_utc = datetime.combine(base, we, tzinfo=IST).astimezone(timezone.utc)
            if tw_start_utc and tw_end_utc <= tw_start_utc:
                # A window whose end reads earlier than its start crosses midnight.
                tw_end_utc += timedelta(days=1)
                issues.append(Issue(row_no, "Tide_Window_End", "INFO", "window_crosses_midnight",
                                    "window end is before its start; treated as next day"))
    elif got.get("tide_window_start_raw") is not None:
        issues.append(Issue(row_no, "Tide_Window_Start", "WARN", "no_anchor_date",
                            "tide window given but no ATA/ETA/Tide_Window_Date to anchor it to"))

    # --- channel depth ------------------------------------------------------
    channel_depth_m = got.get("channel_depth_m")
    depth_source = "COLUMN_Channel_Depth_m"
    if channel_depth_m is None:
        channel_depth_m = REF_CHANNEL_DEPTH_M
        depth_source = f"DEFAULT_{REF_CHANNEL_DEPTH_M}m"
        issues.append(Issue(row_no, "Channel_Depth_m", "INFO", "defaulted",
                            f"blank; assumed the charted {REF_CHANNEL_DEPTH_M} m"))
    channel_depth_m = float(channel_depth_m)
    net_delta = round(channel_depth_m - REF_CHANNEL_DEPTH_M, 4)

    silt = got.get("siltation_m")
    dredge = got.get("dredging_delta_m")
    if silt is None and dredge is None:
        silt = max(0.0, -net_delta)
        dredge = max(0.0, net_delta)
    else:
        silt = float(silt or 0.0)
        dredge = float(dredge or 0.0)
        depth_source = "COLUMN_Siltation_m/Dredging_Delta_m"
        net_delta = round(dredge - silt, 4)

    # --- tide height --------------------------------------------------------
    measured_tide = got.get("tide_height_m")
    if measured_tide is not None:
        tide_height_m, tide_source = float(measured_tide), "COLUMN_Tide_Height_m"
    elif tide_policy == "fixed" and fixed_tide_m is not None:
        tide_height_m, tide_source = float(fixed_tide_m), f"FIXED_{fixed_tide_m}m"
    elif tide_policy == "column":
        tide_height_m, tide_source = 0.0, "COLUMN_MISSING"
        issues.append(Issue(row_no, "Tide_Height_m", "ERROR", "tide_column_required",
                            "--tide-policy column was requested but Tide_Height_m is blank"))
    else:
        when = ata_utc or eta_utc or datetime.now(timezone.utc)
        tide_height_m, tide_source = round(harmonic_tide_m(when), 3), "SYNTHETIC_HARMONIC_v1"

    # --- resources ----------------------------------------------------------
    def availability(field_name: str, mapping: Mapping[str, int], roster: int, column: str) -> int:
        raw_value = got.get(field_name)
        if _is_blank(raw_value):
            return roster
        try:
            return int(parse_choice(raw_value, mapping, label=column) or 0)
        except ParseError as exc:
            issues.append(Issue(row_no, column, "WARN", exc.code,
                                f"{exc}; assumed full availability", str(raw_value)))
            return roster

    pilots_available = availability("pilot_available_raw", PILOT_AVAILABILITY_MAP,
                                    PILOT_ROSTER_N, "Pilot_Available")
    tugs_available = availability("tug_available_raw", TUG_AVAILABILITY_MAP,
                                  TUG_ROSTER_N, "Tug_Available")

    # --- weather ------------------------------------------------------------
    weather_raw = str(got.get("weather_raw") or "")
    try:
        weather_severity = int(parse_choice(weather_raw, WEATHER_SEVERITY_MAP,
                                            label="Weather") or 0) if weather_raw else 0
    except ParseError as exc:
        issues.append(Issue(row_no, "Weather", "WARN", exc.code,
                            f"{exc}; assumed severity 0", weather_raw))
        weather_severity = 0

    wind_kn = float(got.get("wind_kn") or 0.0)
    rain_mm_hr = float(got.get("rain_mm_hr") or 0.0)
    severe_flag = int(
        weather_severity >= 2
        or wind_kn >= SEVERE_WEATHER_WIND_KN
        or rain_mm_hr >= SEVERE_WEATHER_RAIN_MMHR
    )

    incident_raw = str(got.get("incident_raw") or "")
    try:
        incident_severity = int(parse_choice(incident_raw, INCIDENT_SEVERITY_MAP,
                                             label="Incident") or 0) if incident_raw else 0
    except ParseError as exc:
        issues.append(Issue(row_no, "Incident", "WARN", exc.code,
                            f"{exc}; assumed severity 0", incident_raw))
        incident_severity = 0

    # --- port state ---------------------------------------------------------
    occupancy = float(got.get("berth_occupancy_pct") or 0.0)
    queue = got.get("anchorage_queue_count")
    if queue is not None:
        queue_source = "COLUMN_Anchorage_Queue"
        anchorage_queue = int(queue)
    else:
        queue_source = "DERIVED_FROM_OCCUPANCY"
        anchorage_queue = max(0, int(round(
            (occupancy - QUEUE_FROM_OCCUPANCY_BASE_PCT) / QUEUE_FROM_OCCUPANCY_PCT_PER_VESSEL
        ))) if occupancy > 0 else 0

    # --- berth --------------------------------------------------------------
    terminal = canonical_terminal(got.get("terminal"))
    berth = canonical_berth(got.get("requested_berth"), terminal)
    max_draft = got.get("terminal_max_draft_m")
    if max_draft is None:
        looked_up = _terminal_max_draft_lookup(berth, terminal)
        if looked_up is None:
            looked_up = 16.5
            issues.append(Issue(row_no, "Requested_Berth", "WARN", "berth_not_in_roster",
                                f"{berth or '(blank)'} is not in the JNPA roster; "
                                f"assumed max draft 16.5 m", str(got.get("requested_berth"))))
        max_draft = looked_up
    max_draft = float(max_draft)
    if draft_m and max_draft and draft_m > max_draft:
        issues.append(Issue(row_no, "Draft_m", "WARN", "draft_exceeds_berth",
                            f"draft {draft_m} m exceeds berth max draft {max_draft} m"))

    distance = got.get("distance_nm")
    distance_source = "COLUMN_Distance_NM" if distance is not None else "DEFAULT"

    return VesselCallInput(
        row=row_no,
        call_id=call_id,
        source_file=source_file,
        vessel_name=vessel_name,
        vessel_id=vessel_id,
        imo=imo,
        voyage=voyage,
        eta_utc=eta_utc,
        ata_utc=ata_utc,
        berth_ready_utc=got.get("berth_ready_ist"),
        tide_window_start_utc=tw_start_utc,
        tide_window_end_utc=tw_end_utc,
        import_teu=import_teu,
        export_teu=export_teu,
        total_teu=total_teu,
        cargo_weight_mt=float(got.get("cargo_weight_mt") or 0.0),
        cargo_type=cargo_type,
        cargo_class=cargo_class,
        draft_m=draft_m,
        loa_m=loa_m,
        vessel_class=vessel_class,
        transit_speed_kn=transit_speed_kn,
        has_bow_thruster=bool(bow_thruster),
        terminal=terminal,
        requested_berth=berth,
        terminal_max_draft_m=max_draft,
        service_hours=float(got.get("service_hours") or DEFAULT_SERVICE_HOURS),
        priority=int(got.get("priority") or DEFAULT_PRIORITY),
        pilots_available=pilots_available,
        pilots_down=max(0, PILOT_ROSTER_N - pilots_available),
        tugs_available=tugs_available,
        tugs_down=max(0, TUG_ROSTER_N - tugs_available),
        cranes_available=int(got.get("cranes_available") or 0),
        weather_raw=weather_raw,
        weather_severity=weather_severity,
        severe_weather_flag=severe_flag,
        wind_kn=wind_kn,
        rain_mm_hr=rain_mm_hr,
        tide_height_m=tide_height_m,
        channel_depth_m=channel_depth_m,
        net_channel_depth_delta_m=net_delta,
        siltation_m=float(silt),
        dredging_delta_m=float(dredge),
        berth_occupancy_pct=occupancy,
        anchorage_queue_count=anchorage_queue,
        incident_raw=incident_raw,
        incident_severity=incident_severity,
        distance_nm=float(distance if distance is not None else DEFAULT_DISTANCE_NM),
        dukc_status_reported=str(got.get("dukc_status_reported") or "").strip().upper(),
        tide_source=tide_source,
        queue_source=queue_source,
        depth_source=depth_source,
        speed_source=speed_source,
        distance_source=distance_source,
        raw=dict(values),
        issues=issues,
    )


def _derive_batch_features(rows: Sequence[VesselCallInput]) -> None:
    """
    Fill the two features that are properties of the *batch*, not of one row.

    ``calls_prev_24h`` counts calls arriving strictly before this one within the
    preceding 24 h -- strict ``<`` so a call never counts itself, matching M3's
    definition exactly. ``extra_arrivals_24h`` is the excess over the nominal
    10 calls/day.
    """
    dated = sorted((r for r in rows if r.ata_utc is not None), key=lambda r: (r.ata_utc, r.call_id))
    for i, r in enumerate(dated):
        cutoff = r.ata_utc - timedelta(hours=24)
        r.calls_prev_24h = sum(
            1 for other in dated[:i] if other.ata_utc is not None and other.ata_utc >= cutoff
        )
        r.extra_arrivals_24h = max(0, r.calls_prev_24h - NOMINAL_CALLS_PER_DAY)


def load_input(
    path: str,
    *,
    sheet: Optional[str] = None,
    tide_policy: str = "harmonic",
    fixed_tide_m: Optional[float] = None,
) -> InputBatch:
    """
    Read, validate and normalise an input file.

    ``tide_policy`` is one of ``harmonic`` (synthetic model, the default),
    ``column`` (require a ``Tide_Height_m`` column) or ``fixed`` (use
    ``fixed_tide_m`` for every row). A ``Tide_Height_m`` column always wins when
    present, whatever the policy.
    """
    if tide_policy not in ("harmonic", "column", "fixed"):
        raise ValueError(f"unknown tide_policy {tide_policy!r}")
    if tide_policy == "fixed" and fixed_tide_m is None:
        raise ValueError("tide_policy='fixed' requires fixed_tide_m")

    fmt, sheet_name, grid = read_table(path, sheet)
    issues: List[Issue] = []
    if not grid:
        raise ValueError(f"{path}: file is empty")

    header_idx = _find_header_row(grid)
    headers = list(grid[header_idx])
    header_map: Dict[str, str] = {}
    unknown: List[str] = []
    col_to_field: Dict[int, str] = {}

    for col_idx, raw_header in enumerate(headers):
        if _is_blank(raw_header):
            continue
        k = _key(raw_header)
        spec = _MATCH_INDEX.get(k)
        if spec is None:
            unknown.append(str(raw_header))
            severity, code, msg = "INFO", "unknown_column", "column not in the schema; carried through as context"
            if k in _TARGET_KEYS:
                severity, code = "ERROR", "target_column_in_input"
                msg = ("this is a TARGET column (ETB/TAT/ETD or an actual outcome). "
                       "Remove it from the input file -- the models predict it")
            issues.append(Issue(0, str(raw_header), severity, code, msg))
            continue
        if spec.field_name in col_to_field.values():
            issues.append(Issue(0, str(raw_header), "WARN", "duplicate_column",
                                f"maps to {spec.canonical}, which is already bound; ignored"))
            continue
        col_to_field[col_idx] = spec.field_name
        header_map[str(raw_header)] = spec.canonical

    missing = tuple(
        s.canonical for s in COLUMN_SPECS
        if s.required and s.field_name not in col_to_field.values()
    )
    for name in missing:
        issues.append(Issue(0, name, "ERROR", "missing_required_column",
                            "required column is absent from the file"))

    rows: List[VesselCallInput] = []
    for offset, raw_row in enumerate(grid[header_idx + 1:], start=1):
        if all(_is_blank(c) for c in raw_row):
            continue
        values = {
            field_name: (raw_row[col_idx] if col_idx < len(raw_row) else None)
            for col_idx, field_name in col_to_field.items()
        }
        extras = {
            str(headers[i]): raw_row[i]
            for i in range(min(len(headers), len(raw_row)))
            if i not in col_to_field and not _is_blank(headers[i])
        }
        merged = dict(values)
        merged.update({f"_extra_{k}": v for k, v in extras.items()})
        rows.append(_parse_row(
            header_idx + 1 + offset, values,
            tide_policy=tide_policy, fixed_tide_m=fixed_tide_m,
            source_file=os.path.basename(path),
        ))
        rows[-1].raw = merged

    _derive_batch_features(rows)

    if not rows:
        issues.append(Issue(0, "", "ERROR", "no_data_rows",
                            "the file has a header but no data rows"))

    return InputBatch(
        schema_version=SCHEMA_VERSION,
        source_file=os.path.abspath(path),
        source_format=fmt,
        sheet_name=sheet_name,
        header_map=header_map,
        unknown_columns=tuple(unknown),
        missing_required=missing,
        rows=rows,
        issues=issues,
        read_at_utc=datetime.now(timezone.utc),
    )


# ---------------------------------------------------------------------------
# SECTION 8 -- adapters: VesselCallInput -> each model's dataclasses
# ---------------------------------------------------------------------------


def to_m1_vessel(call: VesselCallInput):
    """-> ``uc1_m1_dukc.VesselState``"""
    import uc1_m1_dukc as m1

    return m1.VesselState(
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        vessel_class=call.cargo_class,   # CONTAINER -> Cb 0.65, BULK -> Cb 0.80
        static_draft_m=call.draft_m,
        transit_speed_kn=call.transit_speed_kn,
        loa_m=call.loa_m or 300.0,
        beam_m=max(20.0, round((call.loa_m or 300.0) / 7.5, 1)),
    )


def to_m1_channel(call: VesselCallInput, reach_id: str = "CH-INNER"):
    """-> ``uc1_m1_dukc.ChannelState`` for one reach."""
    import uc1_m1_dukc as m1

    return m1.ChannelState(
        reach=m1.DEFAULT_REACHES[reach_id],
        tide_height_m=call.tide_height_m,
        siltation_delta_m=call.siltation_m,
        dredging_delta_m=call.dredging_delta_m,
    )


def to_m2_vessel(call: VesselCallInput):
    """-> ``uc1_m2_tidal_window.VesselState``"""
    import uc1_m2_tidal_window as m2

    return m2.VesselState(
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        vessel_class=call.cargo_class,
        static_draft_m=call.draft_m,
        transit_speed_kn=call.transit_speed_kn,
        loa_m=call.loa_m or 300.0,
        beam_m=max(20.0, round((call.loa_m or 300.0) / 7.5, 1)),
    )


def to_m3_features(call: VesselCallInput):
    """
    -> ``uc1_m3_tat_predict.TATFeatures`` (pre-berthing information set only).

    Note what is deliberately absent: no ATD, no TAT, no berth-stay hours, and
    no ``Cranes_Available`` (not in ``FEATURE_COLUMNS`` -- see the module
    docstring). ``atb_utc`` is carried because M3 uses it as the split key, and
    M3's own import-time assertion bans it as a predictor.
    """
    import uc1_m3_tat_predict as m3

    return m3.TATFeatures(
        call_id=call.call_id,
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        terminal=call.terminal,
        berth_id=call.requested_berth,
        atb_utc=call.ata_utc or datetime.now(timezone.utc),
        parcel_teu=int(call.total_teu),
        draft_m=call.draft_m,
        terminal_max_draft_m=call.terminal_max_draft_m,
        draft_vs_terminal_max_m=round(call.draft_m - call.terminal_max_draft_m, 3),
        weather_severity=call.weather_severity,
        severe_weather_flag=call.severe_weather_flag,
        rain_mm_hr=call.rain_mm_hr,
        wind_kn=call.wind_kn,
        net_channel_depth_delta_m=call.net_channel_depth_delta_m,
        pilots_down=call.pilots_down,
        tugs_down=call.tugs_down,
        anchorage_queue_count=call.anchorage_queue_count,
        extra_arrivals_24h=call.extra_arrivals_24h,
        incident_severity=call.incident_severity,
        berth_window_extension_h=call.berth_window_extension_h,
        calls_prev_24h=call.calls_prev_24h,
    )


def to_m4_eta_observation(call: VesselCallInput, now_utc: Optional[datetime] = None,
                          ais_staleness_minutes: float = 15.0):
    """-> ``uc1_m4_berth_utilisation.EtaObservation``"""
    import uc1_m4_berth_utilisation as m4

    eta = call.eta_utc or call.ata_utc or datetime.now(timezone.utc)
    now = now_utc or (eta - timedelta(hours=12))
    return m4.EtaObservation(
        call_id=call.call_id,
        vessel_id=call.vessel_id,
        now_utc=now,
        forecast_eta_utc=eta,
        ais_staleness_minutes=ais_staleness_minutes,
        source="INPUT_FILE",
        vessel_name=call.vessel_name,
    )


def to_m5_request(call: VesselCallInput):
    """-> ``uc1_m5_berth_optimiser.BerthRequest``"""
    import uc1_m5_berth_optimiser as m5

    start = call.ata_utc or call.eta_utc or datetime.now(timezone.utc)
    return m5.BerthRequest(
        request_id=call.call_id,
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        loa_m=call.loa_m or 300.0,
        draft_m=call.draft_m,
        requested_berth_id=call.requested_berth,
        requested_start_utc=start,
        service_hours=call.service_hours,
        priority=call.priority,
        earliest_start_utc=start,
        latest_start_utc=None,
    )


def to_m5_tidal_window(call: VesselCallInput):
    """-> ``uc1_m5_berth_optimiser.TidalWindow`` from the sheet's stated window."""
    import uc1_m5_berth_optimiser as m5

    if call.tide_window_start_utc is None or call.tide_window_end_utc is None:
        return None
    return m5.TidalWindow(
        window_id=f"TW-{call.call_id}",
        start_utc=call.tide_window_start_utc,
        end_utc=call.tide_window_end_utc,
        max_draft_m=max(call.draft_m, call.terminal_max_draft_m),
        reach_id="CH-INNER",
        direction="INBOUND",
    )


def to_m6_vessel(call: VesselCallInput, now_utc: Optional[datetime] = None):
    """-> ``uc1_m6_jit_rta.VesselAtSea``"""
    import uc1_m6_jit_rta as m6

    arrival = call.ata_utc or call.eta_utc or datetime.now(timezone.utc)
    now = now_utc or (arrival - timedelta(hours=call.distance_nm / 16.0))
    return m6.VesselAtSea(
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        vessel_class=call.vessel_class,
        now=now,
        distance_to_go_nm=call.distance_nm,
        arrival_draft_m=call.draft_m,
    )


def to_m6_readiness(call: VesselCallInput, berth_ready_utc: Optional[datetime] = None):
    """-> ``uc1_m6_jit_rta.PortReadiness``"""
    import uc1_m6_jit_rta as m6

    arrival = call.ata_utc or call.eta_utc or datetime.now(timezone.utc)
    ready = berth_ready_utc or call.berth_ready_utc
    if ready is None:
        # No berth-ready column: fall back to the stated tidal window opening,
        # else the arrival itself. Recorded in the breakdown as an assumption.
        ready = call.tide_window_start_utc or arrival
    return m6.PortReadiness(
        berth_id=call.requested_berth,
        berth_ready_time=ready,
        tidal_window_start=call.tide_window_start_utc,
        tidal_window_end=call.tide_window_end_utc,
        tidal_window_max_draft_m=call.terminal_max_draft_m or None,
    )


def to_m7_movement(call: VesselCallInput):
    """-> ``uc1_m7_port_craft.VesselMovement`` (the inbound berthing)."""
    import uc1_m7_port_craft as m7

    start = call.ata_utc or call.eta_utc or datetime.now(timezone.utc)
    req = m7.requirements_for("BERTHING", call.vessel_class, call.has_bow_thruster)
    return m7.VesselMovement(
        movement_id=f"MV-{call.call_id}",
        vessel_id=call.vessel_id,
        vessel_name=call.vessel_name,
        movement_type="BERTHING",
        vessel_class=call.vessel_class,
        berth_id=call.requested_berth,
        start_utc=start,
        end_utc=start + timedelta(hours=req.duration_hours),
        req_pilots=req.pilots,
        req_tugs=req.tugs,
        req_mooring=req.mooring,
        priority=call.priority,
        has_bow_thruster=call.has_bow_thruster,
        tide_locked=call.tide_window_start_utc is not None,
    )


def to_m8_disruptions(call: VesselCallInput) -> List[Any]:
    """
    -> ``list[uc1_m8_causal_chain.Disruption]``.

    Only the graph's exogenous nodes may be set; the rest are computed by
    propagation. Values equal to the node baseline are skipped so the
    disruption list stays honest about what actually changed.
    """
    import uc1_m8_causal_chain as m8

    graph = m8.build_graph()
    baselines = {node_id: node.baseline for node_id, node in graph.nodes.items()}
    proposed: List[Tuple[str, float, str]] = [
        ("WX_WIND_KN", call.wind_kn, f"wind {call.wind_kn:g} kn"),
        ("WX_RAIN_MMHR", call.rain_mm_hr, f"rain {call.rain_mm_hr:g} mm/hr"),
        ("TIDE_HEIGHT_M", call.tide_height_m, f"tide {call.tide_height_m:.2f} m"),
        ("SILTATION_M", call.siltation_m, f"siltation {call.siltation_m:.2f} m"),
        ("DREDGING_DELTA_M", call.dredging_delta_m, f"dredging +{call.dredging_delta_m:.2f} m"),
        ("PILOT_AVAIL_N", float(call.pilots_available), f"{call.pilots_available} pilots"),
        ("TUG_AVAIL_N", float(call.tugs_available), f"{call.tugs_available} tugs"),
        ("ARRIVAL_DEMAND_N", float(max(call.calls_prev_24h, 1)),
         f"{call.calls_prev_24h} arrivals/24 h"),
    ]
    out: List[Any] = []
    for node_id, value, label in proposed:
        if abs(float(value) - float(baselines.get(node_id, 0.0))) < 1e-9:
            continue
        out.append(m8.Disruption(node_id=node_id, value=float(value),
                                 label=label, kind="SENSOR"))
    return out


# ---------------------------------------------------------------------------
# SECTION 9 -- template emitter
# ---------------------------------------------------------------------------


def template_rows() -> Tuple[List[str], List[List[Any]]]:
    """Header + one worked example row, covering required and optional columns."""
    headers = [s.canonical for s in COLUMN_SPECS]
    example = [s.example for s in COLUMN_SPECS]
    return headers, [example]


def emit_template(path: str) -> str:
    """
    Write a blank input template with a documented README sheet.

    Produces .xlsx when openpyxl is available, otherwise .csv alongside a
    ``*_SCHEMA.md``, so the template is always obtainable.
    """
    headers, examples = template_rows()
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm") and _HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Training_Data"
        ws.append(headers)
        for row in examples:
            ws.append(list(row))
        for i, spec in enumerate(COLUMN_SPECS, start=1):
            ws.cell(row=1, column=i).comment = None
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(
                12, min(26, len(spec.canonical) + 4)
            )
        doc = wb.create_sheet("README")
        doc.append(["JNPA UC-1 vessel-call input template", "", "", "", ""])
        doc.append([f"schema {SCHEMA_VERSION}", f"generated by {MODULE_VERSION}"])
        doc.append([])
        doc.append(["Column", "Required", "Type", "Unit", "Meaning"])
        for spec in COLUMN_SPECS:
            doc.append([spec.canonical, "YES" if spec.required else "optional",
                        spec.kind, spec.unit, spec.description])
        doc.append([])
        doc.append(["TARGETS (produced by the models -- do NOT put these in Training_Data):"])
        for t in TARGET_COLUMNS:
            doc.append([t])
        doc.append([])
        doc.append(["All timestamps are IST (UTC+05:30). Blank optional cells are derived."])
        for col, width in (("A", 24), ("B", 12), ("C", 12), ("D", 10), ("E", 80)):
            doc.column_dimensions[col].width = width
        wb.save(path)
        return path

    csv_path = os.path.splitext(path)[0] + ".csv"
    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        for row in examples:
            w.writerow(row)
    md_path = os.path.splitext(path)[0] + "_SCHEMA.md"
    with open(md_path, "w", encoding="utf-8") as fh:
        fh.write(f"# JNPA UC-1 input schema ({SCHEMA_VERSION})\n\n")
        fh.write("| Column | Required | Type | Unit | Meaning |\n")
        fh.write("|---|---|---|---|---|\n")
        for spec in COLUMN_SPECS:
            fh.write(f"| `{spec.canonical}` | {'YES' if spec.required else 'optional'} "
                     f"| {spec.kind} | {spec.unit} | {spec.description} |\n")
        fh.write("\n**Targets (never inputs):** " + ", ".join(f"`{t}`" for t in TARGET_COLUMNS) + "\n")
    return csv_path


# ---------------------------------------------------------------------------
# SECTION 10 -- self-test
# ---------------------------------------------------------------------------


def _self_test() -> List[Tuple[str, bool, str]]:
    checks: List[Tuple[str, bool, str]] = []

    def check(name: str, passed: bool, detail: str = "") -> None:
        checks.append((name, bool(passed), detail))

    # --- parsers ------------------------------------------------------------
    check("parse_float_units", parse_float("13.2 m") == 13.2, "strips a trailing unit")
    check("parse_float_pct", parse_float("68%") == 68.0, "strips %")
    check("parse_float_comma", parse_float("52,000") == 52000.0, "strips thousands separator")
    try:
        parse_float("abc")
        check("parse_float_rejects_text", False, "no exception raised")
    except ParseError as exc:
        check("parse_float_rejects_text", exc.code == "not_a_number", exc.code)

    dt = parse_datetime_ist("2026-07-29 06:00")
    check("ist_to_utc", dt == datetime(2026, 7, 29, 0, 30, tzinfo=timezone.utc),
          f"06:00 IST -> {dt.isoformat()}")
    check("ist_to_utc_ddmmyyyy",
          parse_datetime_ist("29-07-2026 06:00") == datetime(2026, 7, 29, 0, 30, tzinfo=timezone.utc),
          "DD-MM-YYYY accepted")
    check("berman_format",
          parse_datetime_ist("11022026:17:00") == datetime(2026, 2, 11, 11, 30, tzinfo=timezone.utc),
          "BERMAN DDMMYYYY:HH:MM accepted")
    check("parse_time", parse_time_of_day("06:30") == time(6, 30), "HH:MM")

    # --- canonicalisation ---------------------------------------------------
    check("berth_cb04", canonical_berth("CB04") == "CB-04", canonical_berth("CB04"))
    check("berth_bm05", canonical_berth("BM05") == "BMCT-05", canonical_berth("BM05"))
    check("berth_ccb", canonical_berth("CCB-N") == "CCB-N", canonical_berth("CCB-N"))
    check("terminal_jnpct", canonical_terminal("JNPCT") == "NSFT",
          "JNPCT is the former name of NSFT")

    # --- categorical maps ---------------------------------------------------
    check("weather_moderate", parse_choice("Moderate Rain", WEATHER_SEVERITY_MAP, label="w") == 2)
    check("weather_substring", parse_choice("Heavy Rain squalls", WEATHER_SEVERITY_MAP, label="w") == 3,
          "longest-key-first substring match")
    check("incident_minor", parse_choice("Minor Delay", INCIDENT_SEVERITY_MAP, label="i") == 1)
    check("vessel_class_loa", derive_vessel_class(182.08, 3600) == "FEEDER",
          derive_vessel_class(182.08, 3600))
    check("vessel_class_ulcv", derive_vessel_class(399.0, 20000) == "ULCV")
    check("cargo_class_bulk", derive_cargo_class("Iron Ore in bulk") == "BULK")

    # --- schema integrity ---------------------------------------------------
    field_names = [s.field_name for s in COLUMN_SPECS]
    check("no_duplicate_fields", len(field_names) == len(set(field_names)),
          f"{len(field_names)} columns")
    dataclass_fields = {f.name for f in VesselCallInput.__dataclass_fields__.values()}
    unmapped = [
        s.field_name for s in COLUMN_SPECS
        if s.field_name not in dataclass_fields
        and not s.field_name.endswith("_raw")
        and s.field_name not in ("eta_ist", "ata_ist", "berth_ready_ist", "vessel_class_raw")
    ]
    check("all_columns_land_somewhere", not unmapped, f"unmapped: {unmapped}")
    check("targets_are_blocked", all(_key(t) in _TARGET_KEYS for t in TARGET_COLUMNS),
          "ETB/TAT/ETD rejected as input columns")

    # --- round trip through a temporary CSV ---------------------------------
    import tempfile

    headers, examples = template_rows()
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "t.csv")
        with open(csv_path, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(headers)
            for row in examples:
                w.writerow(["" if v == "" else v for v in row])
        batch = load_input(csv_path)
        check("template_is_valid_input", batch.ok,
              f"{len(batch.valid_rows)} valid, {batch.error_count} errors: "
              + "; ".join(str(i) for i in batch.all_issues if i.severity == "ERROR")[:160])
        if batch.valid_rows:
            row = batch.valid_rows[0]
            check("template_row_draft", abs(row.draft_m - 13.2) < 1e-9, f"{row.draft_m}")
            check("template_row_berth", row.requested_berth == "CB-04", row.requested_berth)
            check("template_row_teu", row.total_teu == 3600, str(row.total_teu))

        # A target column in the input file must be rejected outright.
        bad_path = os.path.join(tmp, "bad.csv")
        with open(bad_path, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["Vessel", "ATA", "Draft_m", "TAT"])
            w.writerow(["X", "2026-07-29 06:00", "13.0", "44"])
        bad = load_input(bad_path)
        check("target_column_rejected",
              any(i.code == "target_column_in_input" for i in bad.all_issues),
              "TAT column in the input file raises an ERROR")

    # --- adapters -----------------------------------------------------------
    sample = VesselCallInput(
        row=2, call_id="C-0002", vessel_name="TEST", vessel_id="V1",
        ata_utc=datetime(2026, 7, 29, 0, 30, tzinfo=timezone.utc),
        draft_m=14.1, loa_m=207.95, total_teu=4800, terminal="NSFT",
        requested_berth="CB-02", terminal_max_draft_m=16.5, tide_height_m=2.6,
        channel_depth_m=14.6, siltation_m=0.4, net_channel_depth_delta_m=-0.4,
        wind_kn=22.0, rain_mm_hr=8.0, weather_severity=2, severe_weather_flag=1,
        pilots_available=2, pilots_down=1, tugs_available=4, tugs_down=0,
        anchorage_queue_count=6, vessel_class="FEEDER", cargo_class="CONTAINER",
    )
    try:
        v1 = to_m1_vessel(sample)
        c1 = to_m1_channel(sample)
        import uc1_m1_dukc as m1

        res = m1.evaluate_dukc(v1, c1, with_sensitivity=False)
        check("adapter_m1", res.status in ("SAFE", "MARGINAL", "NO GO"),
              f"net UKC {res.net_ukc_m:.3f} m -> {res.status}")
    except Exception as exc:  # pragma: no cover
        check("adapter_m1", False, repr(exc)[:150])

    try:
        f3 = to_m3_features(sample)
        import uc1_m3_tat_predict as m3

        vec = m3.to_vector(f3)
        check("adapter_m3", len(vec) == len(m3.FEATURE_COLUMNS),
              f"{len(vec)} features, none banned")
        check("adapter_m3_no_label",
              not any(b in m3.FEATURE_COLUMNS for b in m3.BANNED_FIELDS),
              "feature vector carries no target field")
    except Exception as exc:  # pragma: no cover
        check("adapter_m3", False, repr(exc)[:150])

    for name, fn in (("adapter_m5", to_m5_request), ("adapter_m7", to_m7_movement),
                     ("adapter_m6_vessel", to_m6_vessel), ("adapter_m4", to_m4_eta_observation)):
        try:
            obj = fn(sample)
            check(name, obj is not None, type(obj).__name__)
        except Exception as exc:  # pragma: no cover
            check(name, False, repr(exc)[:150])

    try:
        d8 = to_m8_disruptions(sample)
        check("adapter_m8", len(d8) > 0, f"{len(d8)} exogenous disruptions")
        import uc1_m8_causal_chain as m8

        check("adapter_m8_exogenous_only",
              all(d.node_id in m8._EXOGENOUS for d in d8),
              "only exogenous nodes are perturbed")
    except Exception as exc:  # pragma: no cover
        check("adapter_m8", False, repr(exc)[:150])

    # --- xlsx readers agree -------------------------------------------------
    sample_xlsx = jnpa_paths.SAMPLE_INPUT_XLSX
    if os.path.exists(sample_xlsx):
        try:
            _, _, stdlib_grid = ("", "", None), None, None  # placeholder for clarity
        except Exception:
            pass
        try:
            name_a, grid_a = _read_xlsx_stdlib(sample_xlsx)
            if _HAS_OPENPYXL:
                wb = openpyxl.load_workbook(sample_xlsx, data_only=True, read_only=True)
                ws = wb.worksheets[0]
                grid_b = [list(r) for r in ws.iter_rows(values_only=True)]
                wb.close()
                same_header = [str(x) for x in grid_a[0]] == [str(x) for x in grid_b[0] if x is not None]
                check("xlsx_readers_agree_header", same_header,
                      "stdlib reader matches openpyxl on the header row")
                a_vals = [str(v) for v in grid_a[1][:12]]
                b_vals = [str(v) for v in grid_b[1][:12]]
                check("xlsx_readers_agree_row1", a_vals == b_vals,
                      f"stdlib {a_vals[3]!r} vs openpyxl {b_vals[3]!r}")
            else:  # pragma: no cover
                check("xlsx_readers_agree_header", len(grid_a) > 1, "openpyxl absent; stdlib only")
        except Exception as exc:  # pragma: no cover
            check("xlsx_readers_agree_header", False, repr(exc)[:150])

        batch = load_input(sample_xlsx)
        check("sample_workbook_loads", batch.ok,
              f"{len(batch.valid_rows)}/{len(batch.rows)} rows valid, "
              f"{batch.error_count} errors, {batch.warn_count} warnings")
        check("sample_has_25_columns", len(batch.header_map) >= 25,
              f"{len(batch.header_map)} columns matched")
        if batch.valid_rows:
            r0 = batch.valid_rows[0]
            check("sample_ata_utc",
                  r0.ata_utc == datetime(2026, 7, 29, 2, 15, tzinfo=timezone.utc),
                  f"07:45 IST -> {r0.ata_utc.isoformat()}")
            check("sample_dukc_status_not_a_feature",
                  r0.dukc_status_reported == "SAFE",
                  "carried as a label, never fed to M1")
            check("sample_calls_prev_24h_no_self_count",
                  all(x.calls_prev_24h < len(batch.valid_rows) for x in batch.valid_rows),
                  "strict < on ATA, so a call never counts itself")

    return checks


# ---------------------------------------------------------------------------
# SECTION 11 -- CLI
# ---------------------------------------------------------------------------


def _print_batch(batch: InputBatch, verbose: bool) -> None:
    s = batch.summary()
    print("=" * 78)
    print(f"  JNPA UC-1 INPUT ADAPTER  |  {MODULE_VERSION}  |  schema {SCHEMA_VERSION}")
    print("=" * 78)
    print(f"  file          : {s['source_file']}")
    print(f"  format        : {s['source_format']}   sheet: {s['sheet_name']}")
    print(f"  rows          : {s['rows_valid']} valid / {s['rows_read']} read")
    print(f"  columns       : {s['columns_matched']} matched, "
          f"{len(s['unknown_columns'])} unknown")
    if s["unknown_columns"]:
        print(f"                  unknown: {', '.join(s['unknown_columns'])}")
    if s["missing_required"]:
        print(f"  MISSING       : {', '.join(s['missing_required'])}")
    print(f"  issues        : {s['errors']} ERROR, {s['warnings']} WARN")
    print()

    if batch.valid_rows:
        print("  NORMALISED ROWS")
        hdr = (f"  {'row':>3}  {'vessel':<24} {'ATA (UTC)':<17} {'draft':>6} "
               f"{'TEU':>6} {'berth':<9} {'tide':>6} {'wind':>5} {'queue':>6} {'class':<13}")
        print(hdr)
        print("  " + "-" * (len(hdr) - 2))
        for r in batch.valid_rows:
            ata = r.ata_utc.strftime("%Y-%m-%d %H:%M") if r.ata_utc else "-"
            print(f"  {r.row:>3}  {r.vessel_name[:24]:<24} {ata:<17} "
                  f"{r.draft_m:>6.2f} {r.total_teu:>6} {r.requested_berth:<9} "
                  f"{r.tide_height_m:>6.2f} {r.wind_kn:>5.0f} "
                  f"{r.anchorage_queue_count:>6} {r.vessel_class:<13}")
        print()
        r = batch.valid_rows[0]
        print("  PROVENANCE (row %d)" % r.row)
        print(f"    tide_source     : {r.tide_source}")
        print(f"    depth_source    : {r.depth_source}   "
              f"(net delta {r.net_channel_depth_delta_m:+.2f} m vs {REF_CHANNEL_DEPTH_M} m)")
        print(f"    queue_source    : {r.queue_source}")
        print(f"    speed_source    : {r.speed_source} ({r.transit_speed_kn:.1f} kn)")
        print(f"    DUKC_Status     : {r.dukc_status_reported or '(blank)'} "
              f"-- LABEL ONLY, never fed to M1")
        print()

    issues = batch.all_issues
    if issues:
        shown = [i for i in issues if i.severity != "INFO" or verbose]
        if shown:
            print(f"  ISSUES ({len(shown)} shown of {len(issues)})")
            for i in sorted(shown, key=lambda x: (x.severity != "ERROR", x.row))[:40]:
                print("    " + str(i))
            if len(shown) > 40:
                print(f"    ... and {len(shown) - 40} more")
            print()


def main(argv: Optional[Sequence[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Read, validate and normalise a JNPA UC-1 vessel-call input file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "examples:\n"
            "  python run.py input --input data/input/Vessel_Training_Input_Sample.xlsx\n"
            "  python run.py input --input data.csv --validate --verbose\n"
            "  python run.py input --input data.xlsx --json > normalised.json\n"
            "  python run.py input --emit-template my_input.xlsx\n"
            "  python run.py input --selftest\n"
        ),
    )
    p.add_argument("--input", "-i", help="input .xlsx / .csv / .json")
    p.add_argument("--sheet", help="worksheet name (default: first sheet)")
    p.add_argument("--tide-policy", choices=("harmonic", "column", "fixed"), default="harmonic",
                   help="where tide height comes from (default: harmonic, SYNTHETIC)")
    p.add_argument("--tide-m", type=float, help="fixed tide height for --tide-policy fixed")
    p.add_argument("--validate", action="store_true",
                   help="validate only; exit 1 if any ERROR is found")
    p.add_argument("--emit-template", metavar="PATH", help="write a blank input template")
    p.add_argument("--json", action="store_true", help="emit normalised rows as JSON")
    p.add_argument("--verbose", "-v", action="store_true", help="show INFO issues too")
    p.add_argument("--selftest", action="store_true", help="run the built-in checks")
    args = p.parse_args(argv)

    if args.emit_template:
        written = emit_template(args.emit_template)
        print(f"template written: {written}")
        if not written.endswith(".xlsx"):
            print("  (openpyxl not installed, so CSV + a schema markdown were written instead)")
        return 0

    if args.selftest or not args.input:
        checks = _self_test()
        passed = sum(1 for _, ok, _ in checks if ok)
        print("=" * 78)
        print(f"  jnpa_input.py self-test  |  {MODULE_VERSION}")
        print("=" * 78)
        for name, ok, detail in checks:
            print(f"  [{'PASS' if ok else 'FAIL'}] {name:<34} {detail}")
        print("-" * 78)
        print(f"  {passed}/{len(checks)} checks passed")
        if not args.input:
            print("\n  no --input given; run with --input <file> to normalise a data file")
        return 0 if passed == len(checks) else 1

    try:
        batch = load_input(args.input, sheet=args.sheet,
                           tide_policy=args.tide_policy, fixed_tide_m=args.tide_m)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps({
            "summary": batch.summary(),
            "rows": [r.as_dict() for r in batch.rows],
            "issues": [i.as_dict() for i in batch.issues],
        }, indent=2, default=str))
    else:
        _print_batch(batch, args.verbose)

    if args.validate:
        verdict = "PASS" if batch.ok else "FAIL"
        print(f"  VALIDATION: {verdict}  "
              f"({batch.error_count} errors, {batch.warn_count} warnings)")
    return 0 if batch.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
