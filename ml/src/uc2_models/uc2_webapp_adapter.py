"""
UC2-ADAPTER -- JNPA journey-sheet ingest for the seven UC-II models
===================================================================

Jawaharlal Nehru Port Authority (JNPA) -- Workstream 2, UC-II Improved Cargo
Handling & Logistics Optimization. Tender ref GeM/2026/B/7297343.

WHY THIS MODULE EXISTS
----------------------
The seven UC-II models take small, validated numeric contracts -- ``stream_idx``
0-6, ``siding`` 0/1, ``cto_idx`` 0-3. The web app does not have those. What it
has is a row shaped like ``Cargo_Training_Input_Sample.xlsx``: sixty columns of
PCS-native text -- ``Siding: "T2"``, ``CTO_Index: "CTO-2"``, ``Direction:
"Inbound"``, ``Shipping_Line_Code: "CHZ"``, ``Arrival_DateTime: "2026-06-07 08:27"``.

Putting that translation in the frontend would mean every consumer reimplements
it, and the first one to map ``"CTO-2"`` to index 2 instead of 1 ships a wrong
number that nothing catches. So the translation lives here, once, versioned,
and every response says exactly what it derived and what it assumed.

WHAT IT GUARANTEES
------------------
1. NOTHING IS INVENTED. Where a required model input is absent from the sheet,
   the adapter uses a NAMED default, sets ``degraded: true``, and lists the
   assumption in ``mapping.assumptions[]``. It never silently picks a value.

2. EVERY DERIVATION IS SHOWN. ``mapping.derived[]`` carries one entry per model
   input: the source column, the raw value, the mapped value and the rule that
   did it. A port manager can check the translation by hand.

3. UNRECOGNISED CODES FALL BACK LOUDLY. An unknown shipping line maps to
   ``OTHER`` and says so; it does not map to index 0 (MSC) and pretend.

4. THE MODELS ARE UNCHANGED. This is a pure translation layer in front of the
   existing predict functions. Their contracts, coefficients and validation are
   untouched, so the audited numbers stay audited.

SHEET COVERAGE
--------------
    J1_Import_Lifecycle       -> M1 dwell
    J2_Export_Lifecycle       -> M1 dwell (export streams)
    J3_Gate_Transaction       -> M3 queue point, M6 demand
    J4_Gate_Queue_Forecast    -> M3 queue point and curve  (richest: has lags)
    J5_RMS_Scan_Customs       -> M4 (customs-hold evidence)
    J6_CFS_ECY_Chain          -> M1 dwell (off-dock), M4 chain anomalies
    J7_Rail_Rake              -> M2 rake TAT
    J8_Yard_Empty_Pendency    -> M7 empty pool + reefer plugs, M1 facility_load
    J9_Event_Anomaly          -> M4 (flag-shaped; see FLAG_FINDINGS)
    J10_Vessel_Call_Berthing  -> M5 berth stay, M6/M7 reefer + hazmat counts

USAGE
-----
    python uc2_webapp_adapter.py                 # map + predict every sample row
    python uc2_webapp_adapter.py --sheet J1      # one sheet
    python uc2_webapp_adapter.py --json
    python uc2_webapp_adapter.py --selftest

    from uc2_webapp_adapter import predict_j1_dwell
    out = predict_j1_dwell({"Container_No": "DPWU9011100", ...})
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Sequence, Tuple

for _extra in (os.path.dirname(os.path.abspath(__file__)),
               os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "pipeline")):
    if _extra not in sys.path:
        sys.path.append(_extra)

import uc2_m1_container_dwell as m1        # noqa: E402
import uc2_m2_rake_tat as m2               # noqa: E402
import uc2_m3_gate_queue as m3             # noqa: E402
import uc2_m4_event_anomaly as m4          # noqa: E402
import uc2_m5_discharge_berth_stay as m5   # noqa: E402
import uc2_m6_lane_assignment as m6        # noqa: E402
import uc2_m7_empty_pool_reefer as m7      # noqa: E402

# One row exactly as it appears in a Cargo_Training_Input_Sample.xlsx sheet:
# PCS-native column names, values in whatever shape the upstream system wrote
# them. Declared at module scope because Pydantic v2 resolves endpoint
# annotations against module globals, not against build_router()'s locals.
JnpaRow = Dict[str, Any]

MODULE_ID: str = "UC2-ADAPTER"
MODULE_NAME: str = "JNPA journey-sheet adapter"
MODULE_VERSION: str = "adapter-v1.0.0"
ROUTER_PREFIX: str = "/uc2/webapp"


# ==========================================================================
# SECTION 1 -- VERSIONED CODE TABLES
#
# Every table below is a published mapping, not a heuristic. They are served
# at GET /uc2/webapp/mappings so the frontend can render dropdowns from the
# same source of truth the adapter validates against.
# ==========================================================================

# Shipping line: PCS line/agent codes seen in the JNPA corpus -> M1 LINES index.
# LINES = ("MSC", "MAERSK", "ONE", "CMA_CGM", "HAPAG", "OTHER")
# A code that is not listed maps to OTHER (5) and is reported, never guessed.
LINE_CODE_TO_IDX: Dict[str, int] = {
    # MSC
    "MSC": 0, "MSCU": 0, "MEDU": 0, "MED": 0,
    # Maersk / Hamburg Sud / Safmarine
    "MAERSK": 1, "MSK": 1, "MRK": 1, "MAEU": 1, "SAF": 1, "SUD": 1,
    # Ocean Network Express (NYK / MOL / K-Line)
    "ONE": 2, "ONEU": 2, "NYK": 2, "MOL": 2, "KLN": 2,
    # CMA CGM / APL / ANL
    "CMA": 3, "CMAU": 3, "APL": 3, "ANL": 3, "CGM": 3,
    # Hapag-Lloyd / UASC
    "HAPAG": 4, "HLC": 4, "HLCU": 4, "UAS": 4,
}

# Trade stream. The sheet does not carry M1's stream directly; it carries the
# ingredients. resolve_stream() combines them in a documented priority order.
# STREAMS = (IMPORT_CFS, IMPORT_ICD, IMPORT_DPD, EXPORT_CFS, EXPORT_ICD,
#            TRANSSHIPMENT, EMPTY_RETURN)
STREAM_IMPORT_CFS, STREAM_IMPORT_ICD, STREAM_IMPORT_DPD = 0, 1, 2
STREAM_EXPORT_CFS, STREAM_EXPORT_ICD, STREAM_TRANSSHIPMENT, STREAM_EMPTY = 3, 4, 5, 6

# JNPA PCS Delivery_Mode single-letter codes. 'G' dominates the sample.
DELIVERY_MODE_MEANING: Dict[str, str] = {
    "G": "General / godown delivery -- routed through a CFS",
    "C": "CFS delivery",
    "D": "Direct Port Delivery (DPD)",
    "F": "Factory / direct delivery, treated as DPD for evacuation purposes",
    "R": "Rail / ICD onward movement",
}

# Rail siding. The sheet writes the siding NAME; M2 takes the index.
SIDING_TO_IDX: Dict[str, int] = {"T1": 0, "T2": 1, "0": 0, "1": 1}

# Container Train Operator. The sheet writes "CTO-1".."CTO-4" (1-based) or the
# operator name; M2 takes a 0-based index into CTOS.
# CTOS = ("CONCOR", "GATEWAY", "ADANI", "OTHER_CTO")
CTO_TO_IDX: Dict[str, int] = {
    "CTO-1": 0, "CTO1": 0, "CONCOR": 0, "CCLI": 0,
    "CTO-2": 1, "CTO2": 1, "GATEWAY": 1, "GRFL": 1, "GATEWAY RAIL": 1,
    "CTO-3": 2, "CTO3": 2, "ADANI": 2, "ALL": 2,
    "CTO-4": 3, "CTO4": 3, "OTHER": 3, "OTHER_CTO": 3,
}

# Gate move type -> M6 movement class. Drives lane demand.
MOVE_TYPE_TO_CLASS: Dict[str, str] = {
    "DELIVER IMPORT": "IMPORT_LADEN",
    "GATE OUT IMPORT": "IMPORT_LADEN",
    "RECEIVE IMPORT": "IMPORT_LADEN",
    "GATE IN EXPORT": "EXPORT_LADEN",
    "RECEIVE EXPORT": "EXPORT_LADEN",
    "DELIVER EXPORT": "EXPORT_LADEN",
    "EMPTY IN": "EMPTY",
    "EMPTY OUT": "EMPTY",
    "GATE IN EMPTY": "EMPTY",
    "GATE OUT EMPTY": "EMPTY",
}

# Cargo text -> reefer sensitivity class for M7. Checked longest-first.
CARGO_TEXT_TO_SENSITIVITY: Tuple[Tuple[str, str], ...] = (
    ("AMBIENT CONTROLLED", "AMBIENT_CONTROLLED"),
    ("PHARMACEUTICAL", "PHARMA"),
    ("PHARMA", "PHARMA"),
    ("VACCINE", "PHARMA"),
    ("MEDICINE", "PHARMA"),
    ("FROZEN", "FROZEN"),
    ("DEEP FREEZE", "FROZEN"),
    ("ICE CREAM", "FROZEN"),
    ("CHILLED", "CHILLED"),
    ("FRESH", "CHILLED"),
    ("PERISHABLE", "CHILLED"),
    ("REEFER", "UNKNOWN"),
)

# Text that marks a box as refrigerated when it appears in cargo description.
REEFER_TEXT_MARKERS: Tuple[str, ...] = (
    "REEFER", "FROZEN", "CHILLED", "PHARMA", "PERISHABLE", "TEMPERATURE",
    "COLD", "FRESH", "VACCINE",
)

# Customs states that mean the box is held and will not move.
CUSTOMS_HELD_STATES: Tuple[str, ...] = (
    "HELD", "HOLD", "UNDER_INSPECTION", "UNDER INSPECTION", "SEIZED",
    "DETAINED", "QUERY", "EXAMINATION",
)

# J9 carries pre-computed anomaly flags from an upstream detector. They are a
# DIFFERENT taxonomy from M4's R1-R6 trail rules, so they are passed through
# under their own rule IDs rather than being dressed up as M4 findings.
FLAG_FINDINGS: Dict[str, Dict[str, str]] = {
    "ISO_Code_Mismatch_Flag": {
        "ruleId": "J9-F1", "type": "ANOMALY_ISO_MISMATCH", "severity": "WARN",
        "reason": "ISO size/type disagrees between the manifest and the gate move.",
    },
    "Doc_Field_Mismatch_Flag": {
        "ruleId": "J9-F2", "type": "ANOMALY_DOC_FIELD_MISMATCH", "severity": "WARN",
        "reason": "A document field disagrees with the same field on a linked document.",
    },
    "Duplicate_Doc_Flag": {
        "ruleId": "J9-F3", "type": "ANOMALY_DUPLICATE_DOC", "severity": "WARN",
        "reason": "The same physical document was ingested more than once.",
    },
    "Missing_Upstream_Event_Flag": {
        "ruleId": "J9-F4", "type": "ANOMALY_MISSING_UPSTREAM", "severity": "CRIT",
        "reason": "A required earlier event in the lifecycle was never recorded.",
    },
    "Out_Of_Sequence_Flag": {
        "ruleId": "J9-F5", "type": "ANOMALY_OUT_OF_SEQUENCE", "severity": "WARN",
        "reason": "Events arrived in an order the lifecycle does not allow.",
    },
    "Numeric_Field_Non_Numeric_Flag": {
        "ruleId": "J9-F6", "type": "ANOMALY_DATA_QUALITY", "severity": "CRIT",
        "reason": "A field declared numeric carries non-numeric text.",
    },
    "Truncated_Message_Flag": {
        "ruleId": "J9-F7", "type": "ANOMALY_TRUNCATED_MESSAGE", "severity": "CRIT",
        "reason": "The EDI message was cut short; items after the cut were lost.",
    },
}

# Z-score beyond which a dwell or queue observation is called an outlier.
ZSCORE_OUTLIER_THRESHOLD: float = 2.0

# Named defaults used when the sheet cannot supply a required model input.
# Every use of one of these sets degraded=true and names it in assumptions[].
DEFAULT_ARRIVAL_CADENCE_H: float = 6.0
DEFAULT_FACILITY_LOAD: float = 0.70
DEFAULT_TERMINAL_COUNT: int = 1
DEFAULT_CRANES: float = 3.0
DEFAULT_SERVICE_CAPACITY_PER_HOUR: float = 3.0


# ==========================================================================
# SECTION 2 -- TOLERANT SCALAR READERS
#
# The sheet writes "Yes"/"No", "12", 12, "12.0", "" and None for the same
# concept depending on which upstream system produced the row. These readers
# accept all of it and return None -- never a silent zero -- when absent.
# ==========================================================================

_TRUE_TOKENS = {"Y", "YES", "TRUE", "1", "T"}
_FALSE_TOKENS = {"N", "NO", "FALSE", "0", "F", "NONE", "NIL"}


def _norm(value: Any) -> Optional[str]:
    """Uppercase, trimmed text -- or None for anything blank."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() in {"NA", "N/A", "NULL", "-", "--"}:
        return None
    return text.upper()


def read_bool(row: Dict[str, Any], *keys: str) -> Optional[bool]:
    """First key that carries a recognisable boolean wins."""
    for key in keys:
        token = _norm(row.get(key))
        if token is None:
            continue
        if token in _TRUE_TOKENS:
            return True
        if token in _FALSE_TOKENS:
            return False
    return None


def read_float(row: Dict[str, Any], *keys: str) -> Optional[float]:
    """First key that parses as a finite number wins."""
    for key in keys:
        raw = row.get(key)
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            continue
        try:
            value = float(str(raw).strip().replace(",", ""))
        except (TypeError, ValueError):
            continue
        if math.isfinite(value):
            return value
    return None


def read_int(row: Dict[str, Any], *keys: str) -> Optional[int]:
    value = read_float(row, *keys)
    return None if value is None else int(round(value))


def read_text(row: Dict[str, Any], *keys: str) -> Optional[str]:
    for key in keys:
        raw = row.get(key)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            return text
    return None


def read_timestamp(row: Dict[str, Any], *keys: str) -> Optional[datetime]:
    """
    Parse the sheet's timestamp shapes into an aware UTC datetime.

    The README says timestamps are normalised to 'YYYY-MM-DD HH:MM', but rows
    that came through an EDI path can still carry 'DDMMYYYY:HH:MM' and the
    printed forms carry 'DD/MM/YYYY HH:MM'. All three are accepted; anything
    else returns None rather than a guess.
    """
    for key in keys:
        raw = row.get(key)
        if raw is None:
            continue
        if isinstance(raw, datetime):
            return raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)
        text = str(raw).strip()
        if not text:
            continue
        text = text.replace("Z", "+00:00")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                    "%d%m%Y:%H:%M", "%d-%m-%Y %H:%M", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        try:
            parsed = datetime.fromisoformat(text)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def iso_utc(moment: Optional[datetime]) -> Optional[str]:
    if moment is None:
        return None
    return moment.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


# ==========================================================================
# SECTION 3 -- THE MAPPING LEDGER
#
# Every adapter call builds one of these. It is returned inside the response
# so the caller can see each model input, where it came from, and whether it
# was observed or assumed.
# ==========================================================================

class MappingLedger:
    """Records how each model input was arrived at."""

    def __init__(self, sheet: str, target_module: str) -> None:
        self.sheet = sheet
        self.target_module = target_module
        self.derived: List[Dict[str, Any]] = []
        self.assumptions: List[str] = []
        self.warnings: List[str] = []

    def observed(self, field: str, value: Any, source_column: str,
                 raw: Any = None, rule: str = "direct") -> Any:
        """Record a model input that came from a column in the row."""
        self.derived.append({
            "model_input": field, "value": value, "source": source_column,
            "raw": None if raw is None else str(raw), "rule": rule,
            "observed": True,
        })
        return value

    def assumed(self, field: str, value: Any, why: str) -> Any:
        """Record a model input the sheet could not supply. Sets degraded."""
        self.derived.append({
            "model_input": field, "value": value, "source": "DEFAULT",
            "raw": None, "rule": why, "observed": False,
        })
        self.assumptions.append(f"{field}={value} -- {why}")
        return value

    def warn(self, message: str) -> None:
        self.warnings.append(message)

    @property
    def degraded(self) -> bool:
        return bool(self.assumptions)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "adapter_version": MODULE_VERSION,
            "sheet": self.sheet,
            "target_module": self.target_module,
            "derived": self.derived,
            "assumptions": self.assumptions,
            "warnings": self.warnings,
            "inputs_observed": sum(1 for d in self.derived if d["observed"]),
            "inputs_assumed": sum(1 for d in self.derived if not d["observed"]),
        }


def _merge_degraded(model_response: Dict[str, Any],
                    ledger: MappingLedger) -> Dict[str, Any]:
    """
    Fold the adapter's ledger into the model's own response.

    ``degraded`` becomes true if EITHER the model degraded or the adapter had
    to assume an input -- a prediction built on an assumed facility load is
    weaker than one built on a measured one, and the badge must say so.
    """
    out = dict(model_response)
    out["degraded"] = bool(out.get("degraded")) or ledger.degraded
    out["mapping"] = ledger.as_dict()
    if ledger.degraded:
        out["decision_path"] = (
            f"{out.get('decision_path', '')} | adapter={MODULE_VERSION} "
            f"| assumed={len(ledger.assumptions)}"
        ).strip(" |")
    else:
        out["decision_path"] = (
            f"{out.get('decision_path', '')} | adapter={MODULE_VERSION} "
            f"| all_inputs_observed"
        ).strip(" |")
    return out


# ==========================================================================
# SECTION 4 -- FIELD RESOLVERS
#
# One function per model input that needs more than a direct read. Each is
# testable on its own and each records its reasoning in the ledger.
# ==========================================================================

def resolve_line_idx(row: Dict[str, Any], ledger: MappingLedger) -> int:
    """
    Shipping line index from the PCS line code, or the container's ISO owner
    prefix when the line code is absent or unrecognised.

    The prefix path reuses M1's own OWNER_PREFIX_TO_LINE table, which was
    measured off the corpus -- so the two routes cannot drift apart.
    """
    code = _norm(read_text(row, "Shipping_Line_Code", "Liner_Code",
                           "Shipping_Line", "Line_Code", "Agent_Code"))
    if code and code in LINE_CODE_TO_IDX:
        idx = LINE_CODE_TO_IDX[code]
        return ledger.observed("line_idx", idx, "Shipping_Line_Code", code,
                               f"LINE_CODE_TO_IDX[{code}] -> {m1.LINES[idx]}")

    container = _norm(read_text(row, "Container_No", "Entity_Ref"))
    if container and len(container) >= 4:
        prefix = container[:4]
        if prefix in m1.OWNER_PREFIX_TO_LINE:
            idx = m1.OWNER_PREFIX_TO_LINE[prefix]
            return ledger.observed(
                "line_idx", idx, "Container_No", container,
                f"ISO 6346 owner prefix {prefix} -> {m1.LINES[idx]}")

    other = m1.LINES.index("OTHER")
    if code:
        ledger.warn(f"shipping line code {code!r} is not in LINE_CODE_TO_IDX; "
                    f"mapped to OTHER rather than guessed")
        return ledger.observed("line_idx", other, "Shipping_Line_Code", code,
                               "unrecognised code -> OTHER")
    return ledger.assumed("line_idx", other,
                          "no Shipping_Line_Code and no recognised container "
                          "prefix; OTHER carries the corpus-average line effect")


def resolve_stream_idx(row: Dict[str, Any], ledger: MappingLedger) -> int:
    """
    Trade stream from the row's direction, delivery mode and emptiness.

    Priority, highest first:
        1. an empty box is EMPTY_RETURN whatever else the row says
        2. Origin_ICD / 'ICD Rail' arrival mode means the ICD stream
        3. DPD_Eligible=Yes or Delivery_Mode in (D, F) means DPD
        4. otherwise CFS, which is what Delivery_Mode 'G' denotes
    Direction (import vs export) is taken from the sheet name's journey, the
    Move_Type, or the Pre_Advice_Type.
    """
    status = _norm(read_text(row, "Container_Status", "Loaded_Empty"))
    advice = _norm(read_text(row, "Pre_Advice_Type", "Move_Type")) or ""
    if status in {"MTY", "EMPTY", "E"} or "EMPTY" in advice:
        return ledger.observed("stream_idx", STREAM_EMPTY, "Container_Status",
                               status or advice, "empty box -> EMPTY_RETURN")

    # Direction
    is_export = False
    direction_src = "Pre_Advice_Type"
    if "EXPORT" in advice:
        is_export = True
    elif "IMPORT" in advice:
        is_export = False
    else:
        booking = read_text(row, "Booking_No", "Vessel_Sailing_DateTime",
                            "POD", "Final_POD")
        igm = read_text(row, "IGM_No", "BE_No", "OOC_No")
        if booking and not igm:
            is_export, direction_src = True, "Booking_No/POD present, no IGM"
        elif igm:
            is_export, direction_src = False, "IGM_No present"
        else:
            ledger.warn("neither Pre_Advice_Type/Move_Type nor IGM/booking "
                        "evidence present; assuming import")
            direction_src = "assumed import"

    # Evacuation path
    if read_text(row, "Origin_ICD") or "ICD" in (_norm(read_text(row, "Arrival_Mode")) or ""):
        idx = STREAM_EXPORT_ICD if is_export else STREAM_IMPORT_ICD
        return ledger.observed("stream_idx", idx, "Origin_ICD/Arrival_Mode",
                               read_text(row, "Origin_ICD", "Arrival_Mode"),
                               f"ICD movement, direction from {direction_src}"
                               f" -> {m1.STREAMS[idx]}")

    dpd = read_bool(row, "DPD_Eligible")
    mode = _norm(read_text(row, "Delivery_Mode"))
    if dpd is True or mode in {"D", "F"}:
        if not is_export:
            return ledger.observed("stream_idx", STREAM_IMPORT_DPD,
                                   "DPD_Eligible/Delivery_Mode",
                                   f"DPD_Eligible={dpd}, Delivery_Mode={mode}",
                                   "direct port delivery -> IMPORT_DPD")

    idx = STREAM_EXPORT_CFS if is_export else STREAM_IMPORT_CFS
    return ledger.observed(
        "stream_idx", idx, "Delivery_Mode", mode,
        f"Delivery_Mode {mode!r} = "
        f"{DELIVERY_MODE_MEANING.get(mode or '', 'unmapped code, treated as CFS')}; "
        f"direction from {direction_src} -> {m1.STREAMS[idx]}")


def resolve_customs_flag(row: Dict[str, Any], ledger: MappingLedger) -> int:
    """
    Held for customs examination?

    Two independent signals: the customs state itself, and RMS scan selection.
    Either one means the box will sit. PENDING is NOT a hold -- it means the
    entry has not been filed yet, which is the normal state at gate-in.
    """
    state = _norm(read_text(row, "Customs_Status"))
    if state and any(h in state for h in CUSTOMS_HELD_STATES):
        return ledger.observed("customs_flag", 1, "Customs_Status", state,
                               "customs state is a hold state")

    scan = read_bool(row, "Selected_Scan", "Selected_Scan_Flag")
    if scan is True:
        return ledger.observed("customs_flag", 1, "Selected_Scan", "Yes",
                               "selected for RMS scan -> treated as a hold")

    if state or scan is not None:
        return ledger.observed("customs_flag", 0,
                               "Customs_Status/Selected_Scan",
                               f"{state}/{scan}", "no hold signal present")
    return ledger.assumed("customs_flag", 0,
                          "neither Customs_Status nor Selected_Scan supplied; "
                          "assuming not held, which is the majority case")


def resolve_reefer(row: Dict[str, Any], ledger: MappingLedger) -> int:
    """
    Refrigerated?

    Explicit flag first, then cargo text. The numeric ISO size-type the sheet
    carries ("4510", "2210") does NOT by itself prove reefer status, so it is
    only consulted in its alphabetic form where the type group is unambiguous.
    """
    explicit = read_bool(row, "Is_Reefer", "Reefer_Flag")
    if explicit is not None:
        return ledger.observed("reefer", int(explicit), "Is_Reefer", explicit,
                               "explicit reefer flag")

    text = " ".join(filter(None, [
        read_text(row, "Nature_Of_Cargo") or "",
        read_text(row, "Goods_Description") or "",
        read_text(row, "Commodity") or "",
    ])).upper()
    if text:
        for marker in REEFER_TEXT_MARKERS:
            if marker in text:
                return ledger.observed("reefer", 1, "Nature_Of_Cargo", text[:60],
                                       f"cargo text contains {marker!r}")

    iso = _norm(read_text(row, "ISO_Size_Type"))
    if iso and len(iso) >= 3 and iso[2] in {"R", "H"} and not iso[2].isdigit():
        return ledger.observed("reefer", 1, "ISO_Size_Type", iso,
                               f"ISO 6346 type group {iso[2]!r} is refrigerated")

    if text or iso:
        return ledger.observed("reefer", 0, "Nature_Of_Cargo/ISO_Size_Type",
                               text[:40] or iso, "no refrigeration marker found")
    return ledger.assumed("reefer", 0,
                          "no Is_Reefer flag and no cargo description; "
                          "assuming dry, which is ~95% of the corpus")


def resolve_facility_load(row: Dict[str, Any], ledger: MappingLedger) -> float:
    """
    Yard occupancy at gate-in, as a fraction.

    J8 carries it pre-normalised as Facility_Load; J1 carries it as a percent.
    Berth occupancy is NOT a substitute -- a full berth says nothing about
    yard slots -- so it is never used here.
    """
    direct = read_float(row, "Facility_Load")
    if direct is not None and 0.0 <= direct <= 1.0:
        return ledger.observed("facility_load", round(direct, 4),
                               "Facility_Load", direct, "already a fraction")

    pct = read_float(row, "Terminal_Yard_Utilization_Pct", "Utilization_Pct",
                     "Yard_Export_Stack_Utilization_Pct",
                     "Facility_Utilization_Pct", "Siding_Occupancy_Pct")
    if pct is not None:
        value = max(0.0, min(1.0, pct / 100.0))
        return ledger.observed("facility_load", round(value, 4),
                               "Terminal_Yard_Utilization_Pct", pct,
                               "percent / 100")

    return ledger.assumed("facility_load", DEFAULT_FACILITY_LOAD,
                          "no yard utilisation column supplied; "
                          "0.70 is the corpus median occupancy")


def resolve_arrival_cadence(row: Dict[str, Any], ledger: MappingLedger) -> float:
    """
    Hours since the previous container arrived at this facility.

    Three routes, in order of directness:
        1. the caller computed it and passed it in
        2. an expected discharge volume over 24 h inverts to a cadence
        3. the named default, with the assumption recorded
    """
    direct = read_float(row, "Arrival_Cadence_H", "arrival_cadence_h",
                        "Hours_Since_Previous_Arrival")
    if direct is not None and 0 < direct <= 48:
        return ledger.observed("arrival_cadence_h", round(direct, 3),
                               "Arrival_Cadence_H", direct, "supplied directly")

    per_day = read_float(row, "Expected_Discharge_Next_24h",
                         "Facility_Daily_In_Count")
    if per_day is not None and per_day > 0:
        cadence = max(0.01, min(48.0, 24.0 / per_day))
        return ledger.observed(
            "arrival_cadence_h", round(cadence, 3),
            "Expected_Discharge_Next_24h", per_day,
            f"24 h / {per_day:g} arrivals = {cadence:.3f} h between boxes")

    return ledger.assumed("arrival_cadence_h", DEFAULT_ARRIVAL_CADENCE_H,
                          "no arrival cadence and no daily volume column; "
                          "6.0 h is the corpus mean inter-arrival gap")


def resolve_sensitivity_mix(rows: Sequence[Dict[str, Any]],
                            ledger: MappingLedger) -> Optional[Dict[str, float]]:
    """
    Reefer cargo sensitivity mix, counted from cargo descriptions.

    Returns None when no row carries usable cargo text, which makes M7 fall
    back to its own corpus mix rather than to a mix invented here.
    """
    counts: Dict[str, int] = {}
    for row in rows:
        text = " ".join(filter(None, [
            read_text(row, "Nature_Of_Cargo") or "",
            read_text(row, "Goods_Description") or "",
        ])).upper()
        if not text:
            continue
        for marker, sensitivity in CARGO_TEXT_TO_SENSITIVITY:
            if marker in text:
                counts[sensitivity] = counts.get(sensitivity, 0) + 1
                break
    if not counts:
        ledger.warn("no cargo descriptions carried a sensitivity marker; "
                    "M7 will use its corpus mix")
        return None
    total = sum(counts.values())
    mix = {k: round(v / total, 4) for k, v in counts.items()}
    ledger.observed("sensitivity_mix", mix, "Nature_Of_Cargo", counts,
                    f"counted {total} descriptions with a sensitivity marker")
    return mix


# ==========================================================================
# SECTION 5 -- J1 / J2 / J6  ->  M1 CONTAINER DWELL
# ==========================================================================

def _m1_model() -> Any:
    """M1's process-wide singleton -- trained once, not once per request."""
    return m1.get_predictor()


def predict_j1_dwell(row: Dict[str, Any], sheet: str = "J1_Import_Lifecycle") -> Dict[str, Any]:
    """
    Map one import/export/off-dock lifecycle row to M1 and predict its dwell.

    Accepts J1_Import_Lifecycle, J2_Export_Lifecycle and J6_CFS_ECY_Chain rows
    -- they carry the same evidence under slightly different column names, and
    every reader above tries all of them.
    """
    ledger = MappingLedger(sheet, "UC2-M1")

    features = m1.DwellFeatures(
        stream_idx=resolve_stream_idx(row, ledger),
        line_idx=resolve_line_idx(row, ledger),
        arrival_cadence_h=resolve_arrival_cadence(row, ledger),
        customs_flag=resolve_customs_flag(row, ledger),
        reefer=resolve_reefer(row, ledger),
        facility_load=resolve_facility_load(row, ledger),
    ).validate()

    gate_in = read_timestamp(row, "Arrival_DateTime", "Gate_In_DateTime",
                             "Event_Timestamp", "First_In_TS",
                             "Expected_Arrival", "Doc_Generated_DateTime")
    if gate_in is not None:
        ledger.observed("gate_in_utc", iso_utc(gate_in), "Arrival_DateTime",
                        read_text(row, "Arrival_DateTime", "Gate_In_DateTime",
                                  "Event_Timestamp"),
                        "parsed to UTC; enables calendar departure times")
    else:
        ledger.warn("no arrival timestamp; response carries hour offsets only, "
                    "no predicted departure clock time")

    prediction = _m1_model().predict(features)
    response = prediction.as_dict()
    response["container"] = read_text(row, "Container_No") or ""
    response["terminal"] = read_text(row, "Terminal_Code", "Facility_Code") or ""

    if gate_in is not None:
        departure = gate_in + timedelta(hours=prediction.p50_hours)
        response["gateInUtc"] = iso_utc(gate_in)
        response["predictedDepartureUtc"] = iso_utc(departure)
        response["predictedDepartureWindowUtc"] = [
            iso_utc(gate_in + timedelta(hours=prediction.p10_hours)),
            iso_utc(gate_in + timedelta(hours=prediction.p90_hours)),
        ]

    return _merge_degraded(response, ledger)


def predict_dwell_batch(rows: Sequence[Dict[str, Any]],
                        sheet: str = "J1_Import_Lifecycle") -> Dict[str, Any]:
    """
    Predict dwell for many lifecycle rows, computing arrival cadence from them.

    WHY THIS BEATS CALLING predict_j1_dwell IN A LOOP.
    ``arrival_cadence_h`` is defined as the hours since the previous container
    arrived AT THE SAME FACILITY. No single row can carry it -- but a batch
    can: sort each terminal's rows by arrival time and the gaps are the
    cadence, measured rather than assumed. Feeding rows one at a time makes
    every response fall back to the 6.0 h default and raise ``degraded``.

    So the web app should POST the whole visible page of containers here, not
    one row per request.
    """
    by_terminal: Dict[str, List[Tuple[datetime, int]]] = {}
    for index, row in enumerate(rows):
        moment = read_timestamp(row, "Arrival_DateTime", "Gate_In_DateTime",
                                "Event_Timestamp", "First_In_TS")
        if moment is None:
            continue
        terminal = (read_text(row, "Terminal_Code", "Facility_Code") or "").upper()
        by_terminal.setdefault(terminal, []).append((moment, index))

    cadence_by_index: Dict[int, float] = {}
    for terminal, entries in by_terminal.items():
        entries.sort()
        for position in range(1, len(entries)):
            gap_h = (entries[position][0] - entries[position - 1][0]).total_seconds() / 3600.0
            if 0 < gap_h <= 48.0:
                cadence_by_index[entries[position][1]] = round(gap_h, 3)

    results: List[Dict[str, Any]] = []
    for index, row in enumerate(rows):
        enriched = dict(row)
        if index in cadence_by_index and row.get("Arrival_Cadence_H") is None:
            enriched["Arrival_Cadence_H"] = cadence_by_index[index]
        results.append(predict_j1_dwell(enriched, sheet=sheet))

    measured = len(cadence_by_index)
    return {
        "count": len(results),
        "results": results,
        "cadence": {
            "measured_rows": measured,
            "assumed_rows": len(results) - measured,
            "method": ("hours between consecutive Arrival_DateTime values at "
                       "the same Terminal_Code"),
            "note": ("The first container at each terminal has no predecessor "
                     "in the batch, so its cadence is the named default and "
                     "its response carries degraded=true. Send a wider time "
                     "window to measure more of them."),
        },
    }


# ==========================================================================
# SECTION 6 -- J7  ->  M2 RAKE TAT
# ==========================================================================

def _m2_model() -> Any:
    return m2.get_forecaster()


def predict_j7_rake(row: Dict[str, Any], engine: str = "handling") -> Dict[str, Any]:
    """
    Map one J7_Rail_Rake row to M2 and forecast the rake's turnaround.

    The three translations that matter here -- and that a frontend would get
    wrong -- are ``Siding: "T2"`` -> 1, ``CTO_Index: "CTO-2"`` -> 1 (the sheet
    is 1-based, the model is 0-based), and ``Direction: "Inbound"`` -> 1.
    """
    ledger = MappingLedger("J7_Rail_Rake", "UC2-M2")

    siding_raw = _norm(read_text(row, "Siding", "Siding_Code")) or ""
    if siding_raw in SIDING_TO_IDX:
        siding = ledger.observed("siding", SIDING_TO_IDX[siding_raw], "Siding",
                                 siding_raw, f"SIDING_TO_IDX[{siding_raw}]")
    else:
        ledger.warn(f"siding {siding_raw!r} not recognised; defaulting to T1")
        siding = ledger.assumed("siding", 0,
                                f"siding {siding_raw!r} is not T1 or T2")

    cto_raw = _norm(read_text(row, "CTO_Index", "CTO", "CTO_Name",
                              "Container_Train_Operator")) or ""
    cto_key = cto_raw.replace(" ", "")
    if cto_key in CTO_TO_IDX:
        cto_idx = ledger.observed(
            "cto_idx", CTO_TO_IDX[cto_key], "CTO_Index", cto_raw,
            f"CTO_TO_IDX[{cto_key}] -> {m2.CTOS[CTO_TO_IDX[cto_key]]} "
            f"(sheet is 1-based, model is 0-based)")
    else:
        other = m2.CTOS.index("OTHER_CTO")
        ledger.warn(f"CTO {cto_raw!r} not in CTO_TO_IDX; mapped to OTHER_CTO")
        cto_idx = ledger.observed("cto_idx", other, "CTO_Index", cto_raw,
                                  "unrecognised operator -> OTHER_CTO")

    wagons = read_int(row, "Wagon_Count", "Wagons")
    if wagons is None:
        wagons = ledger.assumed("wagon_count", 45,
                                "no Wagon_Count; 45 is the corpus modal rake")
    else:
        wagons = ledger.observed("wagon_count", wagons, "Wagon_Count", wagons)

    hour = read_int(row, "Arrival_Hour")
    if hour is None:
        arrival = read_timestamp(row, "Arrival_Timestamp", "ETA", "Vessel_ETA")
        if arrival is not None:
            hour = ledger.observed("arrival_hour", arrival.hour,
                                   "Arrival_Timestamp", iso_utc(arrival),
                                   "hour extracted from the timestamp")
        else:
            hour = ledger.assumed("arrival_hour", 10,
                                  "no Arrival_Hour and no Arrival_Timestamp")
    else:
        hour = ledger.observed("arrival_hour", hour % 24, "Arrival_Hour", hour)

    direction = _norm(read_text(row, "Direction")) or ""
    if direction.startswith("IN"):
        inbound = ledger.observed("inbound", 1, "Direction", direction,
                                  "Inbound -> 1 (arriving loaded to unload)")
    elif direction.startswith("OUT"):
        inbound = ledger.observed("inbound", 0, "Direction", direction,
                                  "Outbound -> 0 (being loaded to leave)")
    else:
        inbound = ledger.assumed("inbound", 1,
                                 f"Direction {direction!r} unrecognised; "
                                 "inbound is the corpus majority")

    containers = read_int(row, "Container_Count", "Containers")
    if containers is None:
        teu = read_float(row, "TEU")
        if teu is not None and teu > 0:
            # A rake is 20 ft and 40 ft boxes; 40 ft counts 2 TEU. The observed
            # JNPA rakes are all-40 ft (180 TEU on 90 boxes, 164 on 82), so TEU/2
            # recovers the count exactly there and errs low on a mixed rake --
            # which is the safe direction for a handling-time estimate.
            containers = int(round(teu / 2.0))
            ledger.observed(
                "container_count", containers, "TEU", teu,
                f"no Container_Count; TEU/2 = {containers} boxes "
                f"(assumes 40 ft, which the observed JNPA rakes are)")
    if containers is not None:
        ledger.observed("container_count", containers, "Container_Count",
                        containers,
                        f"observed fill {containers / max(wagons, 1):.2f} "
                        f"boxes/wagon (corpus fallback ratio is "
                        f"{m2.FALLBACK_COMPOSITION['containers_per_wagon']})")
    else:
        ledger.warn("no Container_Count; M2 will estimate it from wagon_count "
                    "using the corpus boxes-per-wagon ratio")

    terminals = read_int(row, "Terminal_Count", "Destination_Terminal_Count")
    if terminals is None:
        destinations = read_text(row, "Destination_Terminal", "Terminal_Code")
        if destinations:
            distinct = {d.strip().upper() for d in str(destinations).replace("|", ",").split(",") if d.strip()}
            terminals = ledger.observed(
                "terminal_count", max(1, len(distinct)), "Destination_Terminal",
                destinations, f"{len(distinct)} distinct destination terminal(s)")
        else:
            terminals = ledger.assumed(
                "terminal_count", DEFAULT_TERMINAL_COUNT,
                "no Destination_Terminal column; a single-destination rake is "
                "the least-shunting assumption")
    else:
        terminals = ledger.observed("terminal_count", terminals,
                                    "Terminal_Count", terminals)

    features = m2.RakeFeatures(
        siding=siding, cto_idx=cto_idx, wagon_count=wagons,
        arrival_hour=hour, inbound=inbound,
        container_count=containers, terminal_count=max(1, min(8, terminals)),
    ).validate()

    eta = read_timestamp(row, "Arrival_Timestamp", "ETA")
    prediction = _m2_model().predict(features, engine=engine)
    response = prediction.as_dict()
    response["rakeId"] = read_text(row, "Rake_ID") or ""
    response["terminal"] = read_text(row, "Terminal_Code") or ""

    if eta is not None:
        ledger.observed("eta_utc", iso_utc(eta), "Arrival_Timestamp",
                        read_text(row, "Arrival_Timestamp"),
                        "enables calendar milestone times")
        response["etaUtc"] = iso_utc(eta)
        response["etaPlacementUtc"] = iso_utc(eta + timedelta(hours=prediction.eta_placement_h))
        response["etaRemovalUtc"] = iso_utc(eta + timedelta(hours=prediction.eta_removal_h))
        response["departureWindowUtc"] = [
            iso_utc(eta + timedelta(hours=prediction.departure_window_h[0])),
            iso_utc(eta + timedelta(hours=prediction.departure_window_h[1])),
        ]

    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 7 -- J3 / J4  ->  M3 GATE QUEUE
# ==========================================================================

def _m3_model() -> Any:
    return m3.get_forecaster()


def _m3_default_inflow() -> float:
    """
    M3's training stand-in for the UC-III truck count.

    The model was trained with the OTHER facility's prior-hour arrivals in
    this slot. When the caller has no camera feed, the mean of that same
    series is the only defensible substitute -- and the response discloses it.
    """
    model = _m3_model()
    values = [s.uc3_truck_inflow for s in getattr(model, "steps", [])]
    return round(sum(values) / len(values), 3) if values else 2.0


def predict_j4_queue(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map one J4_Gate_Queue_Forecast row to M3 and forecast the next hour.

    J4 is the richest sheet for this model: it carries the queue lags and the
    UC-III truck inflow directly, so nothing has to be derived and nothing is
    assumed. A J3 gate-transaction row also works -- it carries
    ``Queue_Length_At_Arrival``, which becomes lag1.
    """
    ledger = MappingLedger("J4_Gate_Queue_Forecast", "UC2-M3")

    lag1 = read_float(row, "Queue_Lag1", "Queue_Length", "Queue_Length_At_Arrival")
    if lag1 is None:
        lag1 = ledger.assumed("queue_lag1", 4.0,
                              "no Queue_Lag1/Queue_Length column; 4.0 is the "
                              "corpus mean queue")
    else:
        ledger.observed("queue_lag1", lag1,
                        "Queue_Lag1" if row.get("Queue_Lag1") is not None
                        else "Queue_Length", lag1)

    lag2 = read_float(row, "Queue_Lag2")
    if lag2 is None:
        lag2 = ledger.assumed("queue_lag2", lag1,
                              "no Queue_Lag2 column; holding lag1 flat assumes "
                              "a steady queue rather than a trend")
    else:
        ledger.observed("queue_lag2", lag2, "Queue_Lag2", lag2)

    hour = read_int(row, "Hour_Of_Day", "Hour")
    if hour is None:
        ts = read_timestamp(row, "Timestamp", "Truck_In_Time", "Event_Timestamp")
        if ts is not None:
            hour = ledger.observed("hour", ts.hour, "Timestamp", iso_utc(ts),
                                   "hour extracted from the timestamp")
        else:
            hour = ledger.assumed("hour", 9, "no Hour_Of_Day and no Timestamp")
    else:
        hour = ledger.observed("hour", hour % 24, "Hour_Of_Day", hour)

    inflow = read_float(row, "UC3_Truck_Inflow_Per_Hr", "UC3_Truck_Inflow",
                        "Trucks_In_Last_Hour")
    if inflow is None:
        features = m3.QueueFeatures.from_hour(lag1, lag2, hour,
                                              _m3_default_inflow())
        ledger.assumed("uc3_truck_inflow", round(features.uc3_truck_inflow, 3),
                       "no UC3_Truck_Inflow_Per_Hr column; M3's training "
                       "stand-in (the other facility's prior-hour arrivals) "
                       "is used and disclosed")
    else:
        ledger.observed("uc3_truck_inflow", inflow,
                        "UC3_Truck_Inflow_Per_Hr", inflow,
                        "road-side truck count supplied by the caller")
        features = m3.QueueFeatures.from_hour(lag1, lag2, hour, inflow)

    features.validate()
    prediction = _m3_model().predict(features)
    response = prediction.as_dict()
    response["gateId"] = read_text(row, "Gate_ID", "Gate_No") or ""
    response["terminal"] = read_text(row, "Terminal_Code") or ""

    ts = read_timestamp(row, "Timestamp", "Truck_In_Time")
    if ts is not None:
        response["forecastForUtc"] = iso_utc(ts + timedelta(hours=1))

    # Lanes open is a real column and it changes the wait, which the model's
    # own formula assumes is 3. Report the corrected wait beside the model's.
    lanes = read_float(row, "Lanes_Open")
    if lanes is not None and lanes > 0:
        per_truck_min = m3.AVG_TRANSACTION_MINUTES
        response["estimatedWaitMinutesAtObservedLanes"] = round(
            prediction.queue_vehicles * per_truck_min / lanes, 1)
        response["lanesOpen"] = lanes
        ledger.observed("lanes_open", lanes, "Lanes_Open", lanes,
                        "used only to restate the wait; the model's own "
                        "estimatedWaitMinutes assumes 3 lanes")

    return _merge_degraded(response, ledger)


def forecast_j4_curve(row: Dict[str, Any], hours: int = 12) -> Dict[str, Any]:
    """
    Roll a J4 row forward for ``hours`` steps.

    This exists because M3's own ``forecast_curve`` seeds itself from a corpus
    gate ("CFS" / "ECY") and JNPA's gate IDs are "NSICT-G1", "GTI-G1". Seeding
    from the caller's own lags is both correct for JNPA and more honest: the
    curve starts from the queue the operator is actually looking at.

    Each step feeds its own output back in, so the band widens by sqrt(step) --
    exactly as M3's native curve does.
    """
    if not 1 <= hours <= 72:
        raise ValueError("hours must be 1..72")

    ledger = MappingLedger("J4_Gate_Queue_Forecast", "UC2-M3")
    model = _m3_model()

    lag1 = read_float(row, "Queue_Lag1", "Queue_Length", "Queue_Length_At_Arrival")
    if lag1 is None:
        lag1 = ledger.assumed("queue_lag1", 4.0, "no queue column supplied")
    else:
        ledger.observed("queue_lag1", lag1, "Queue_Length", lag1, "curve seed")

    lag2 = read_float(row, "Queue_Lag2")
    if lag2 is None:
        lag2 = ledger.assumed("queue_lag2", lag1, "no Queue_Lag2 column")
    else:
        ledger.observed("queue_lag2", lag2, "Queue_Lag2", lag2, "curve seed")

    inflow = read_float(row, "UC3_Truck_Inflow_Per_Hr", "Trucks_In_Last_Hour")
    if inflow is None:
        inflow = ledger.assumed("uc3_truck_inflow", _m3_default_inflow(),
                                "no UC3 truck inflow; training stand-in used")
    else:
        ledger.observed("uc3_truck_inflow", inflow, "UC3_Truck_Inflow_Per_Hr",
                        inflow, "held constant across the curve")

    start = read_timestamp(row, "Timestamp", "Truck_In_Time")
    if start is None:
        hour0 = read_int(row, "Hour_Of_Day")
        if hour0 is None:
            hour0 = ledger.assumed("hour", 9, "no Timestamp and no Hour_Of_Day")
        start = datetime.now(timezone.utc).replace(
            hour=hour0 % 24, minute=0, second=0, microsecond=0)
        ledger.warn("no Timestamp column; curve timestamps are relative to "
                    "today at the supplied hour")
    else:
        ledger.observed("start_ts", iso_utc(start), "Timestamp", None,
                        "curve origin")

    points: List[Dict[str, Any]] = []
    ts = start
    for step in range(1, hours + 1):
        ts = ts + timedelta(hours=1)
        features = m3.QueueFeatures.from_hour(lag1, lag2, ts.hour, inflow).validate()
        pred = model.predict(features)
        widen = math.sqrt(step)
        centre = pred.p50
        point = {
            "ts": iso_utc(ts),
            "stepAhead": step,
            "queueVehicles": round(pred.queue_vehicles, 3),
            "p10": round(max(0.0, centre - (centre - pred.p10) * widen), 3),
            "p50": round(centre, 3),
            "p90": round(centre + (pred.p90 - centre) * widen, 3),
            "deferralRecommended": pred.deferral_recommended,
            "estimatedWaitMinutes": round(pred.estimated_wait_minutes, 1),
        }
        points.append(point)
        lag2, lag1 = lag1, pred.queue_vehicles

    deferral_windows: List[Dict[str, str]] = []
    open_from: Optional[str] = None
    for point in points:
        if point["deferralRecommended"] and open_from is None:
            open_from = point["ts"]
        elif not point["deferralRecommended"] and open_from is not None:
            deferral_windows.append({"from": open_from, "to": point["ts"]})
            open_from = None
    if open_from is not None:
        deferral_windows.append({"from": open_from, "to": points[-1]["ts"]})

    response = {
        "moduleId": m3.MODULE_ID,
        "model_version": m3.MODULE_VERSION,
        "gateId": read_text(row, "Gate_ID", "Gate_No") or "",
        "terminal": read_text(row, "Terminal_Code") or "",
        "hours": hours,
        "points": points,
        "deferralWindows": deferral_windows,
        "deferralThreshold": m3.DEFERRAL_THRESHOLD,
        "worstQueue": round(max(p["p50"] for p in points), 3),
        "worstWaitMinutes": round(max(p["estimatedWaitMinutes"] for p in points), 1),
        "degraded": False,
        "decision_path": ("engine=hist_gradient_boosting | seed=caller_supplied_lags "
                          "| band_widening=sqrt(step)"),
        "note": ("Each step is fed its own previous output, so uncertainty "
                 "compounds and the band widens by sqrt(step). Do not render a "
                 "constant-width ribbon."),
    }
    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 8 -- J9  ->  M4 EVENT ANOMALY
# ==========================================================================

def evaluate_j9_event(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Turn one J9_Event_Anomaly row into findings.

    IMPORTANT -- TWO DIFFERENT DETECTORS.
    M4's own rules (R1-R6) read a container's EVENT TRAIL and decide whether
    the lifecycle is broken. J9 does not carry a trail; it carries one row per
    event with boolean flags an UPSTREAM detector already raised, in a
    different taxonomy (ISO mismatch, duplicate document, truncated message).

    Dressing those flags up as M4 findings would claim a detection M4 did not
    make. So they are passed through under their own rule IDs (J9-F1..F7),
    ``engine`` says ``j9_flag_passthrough``, and M4's rule engine is left to
    the endpoint that actually receives a trail.

    The two z-score columns ARE evaluated here, because a threshold on a
    supplied z-score is a decision this adapter is making and can defend.
    """
    ledger = MappingLedger("J9_Event_Anomaly", "UC2-M4")
    findings: List[Dict[str, Any]] = []

    event_ts = read_timestamp(row, "Event_Timestamp")
    entity = read_text(row, "Entity_Ref", "Container_No") or ""

    for column, spec in FLAG_FINDINGS.items():
        if read_bool(row, column) is True:
            findings.append({
                "ruleId": spec["ruleId"],
                "type": spec["type"],
                "severity": spec["severity"],
                "reason": spec["reason"],
                "ageHours": None,
                "evidence": {
                    "source_column": column,
                    "event_id": read_text(row, "Event_ID"),
                    "event_type": read_text(row, "Event_Type"),
                    "event_ts": iso_utc(event_ts),
                    "source_system": read_text(row, "Source_System"),
                    "message_type": read_text(row, "Message_Type"),
                },
            })
            ledger.observed(f"flag:{column}", True, column, "Yes",
                            f"raised {spec['ruleId']} {spec['type']}")

    dwell_z = read_float(row, "Dwell_Zscore")
    if dwell_z is not None and abs(dwell_z) >= ZSCORE_OUTLIER_THRESHOLD:
        findings.append({
            "ruleId": "J9-Z1", "type": "ANOMALY_DWELL_OUTLIER", "severity": "INFO",
            "reason": (f"Dwell is {dwell_z:.1f} standard deviations from the "
                       f"facility mean (threshold {ZSCORE_OUTLIER_THRESHOLD})."),
            "ageHours": None,
            "evidence": {"source_column": "Dwell_Zscore", "z": dwell_z,
                         "threshold": ZSCORE_OUTLIER_THRESHOLD},
        })
        ledger.observed("dwell_zscore", dwell_z, "Dwell_Zscore", dwell_z,
                        f"|z| >= {ZSCORE_OUTLIER_THRESHOLD} -> outlier")

    queue_z = read_float(row, "Queue_Zscore")
    if queue_z is not None and abs(queue_z) >= ZSCORE_OUTLIER_THRESHOLD:
        findings.append({
            "ruleId": "J9-Z2", "type": "ANOMALY_VOLUME_SPIKE", "severity": "INFO",
            "reason": (f"Gate queue is {queue_z:.1f} standard deviations above "
                       f"normal for this hour."),
            "ageHours": None,
            "evidence": {"source_column": "Queue_Zscore", "z": queue_z,
                         "threshold": ZSCORE_OUTLIER_THRESHOLD},
        })

    severity_rank = {"CRIT": 3, "WARN": 2, "INFO": 1}
    findings.sort(key=lambda f: -severity_rank.get(f["severity"], 0))
    worst = findings[0]["severity"] if findings else None

    prior = read_int(row, "Entity_Prior_Anomaly_Count")
    if prior:
        ledger.observed("prior_anomaly_count", prior,
                        "Entity_Prior_Anomaly_Count", prior,
                        "repeat offender -- raise this in the exception queue")

    response = {
        "moduleId": m4.MODULE_ID,
        "model_version": m4.MODULE_VERSION,
        "eventId": read_text(row, "Event_ID") or "",
        "entityRef": entity,
        "entityType": read_text(row, "Entity_Type") or "",
        "terminal": read_text(row, "Terminal_Code") or "",
        "eventTimestampUtc": iso_utc(event_ts),
        "clean": not findings,
        "findingCount": len(findings),
        "worstSeverity": worst,
        "findings": findings,
        "priorAnomalyCount": prior or 0,
        "engine": "j9_flag_passthrough",
        "degraded": False,
        "decision_path": ("engine=j9_flag_passthrough | rules=J9-F1..F7,J9-Z1,J9-Z2 "
                          "| note=M4 R1-R6 evaluate an event trail, not these flags"),
        "rulesNote": (
            "These findings come from flags an upstream detector already raised "
            "and from z-score thresholds applied here. They are NOT M4's R1-R6 "
            "trail rules. Send an event trail to /uc2/m4/predict to run those."
        ),
    }
    return _merge_degraded(response, ledger)


def evaluate_container_trail(rows: Sequence[Dict[str, Any]],
                             container: str = "",
                             now: Optional[str] = None) -> Dict[str, Any]:
    """
    Build an M4 event trail from lifecycle rows and run the real R1-R6 rules.

    This is the path that uses M4 as designed. Give it every row the web app
    holds for one container -- J1 lifecycle, J3 gate moves, J6 facility moves --
    and it assembles the chronological trail M4 expects.
    """
    ledger = MappingLedger("J1/J3/J6 lifecycle rows", "UC2-M4")
    trail: List[Dict[str, str]] = []

    # (column carrying the timestamp, event type it implies, guard)
    EVENT_SOURCES: Tuple[Tuple[Tuple[str, ...], str, Optional[str]], ...] = (
        (("Arrival_DateTime", "Gate_In_DateTime", "First_In_TS"), "GATE_IN", None),
        (("CFS_In_TS",), "GATE_IN", None),
        (("Gate_Out_DateTime", "Last_Out_TS", "CFS_Out_TS", "ECY_Out_TS"), "GATE_OUT", None),
        (("Entry_Inward",), "CUSTOMS_FLAG", "Selected_Scan"),
        (("Scan_Start_TS",), "SCAN_START", None),
        (("Scanner_Stamp", "Scan_Done_TS"), "SCAN_DONE", None),
        (("OOC_DateTime",), "LEO", None),
        (("EDO_Issued_DateTime",), "LEO", None),
    )

    for row in rows:
        if not container:
            container = read_text(row, "Container_No", "Entity_Ref") or ""
        for columns, event_type, guard in EVENT_SOURCES:
            if guard is not None and read_bool(row, guard) is not True:
                continue
            moment = read_timestamp(row, *columns)
            if moment is not None:
                trail.append({"eventType": event_type, "ts": iso_utc(moment)})
                ledger.observed(f"event:{event_type}", iso_utc(moment),
                                "/".join(columns), None,
                                f"lifecycle column -> {event_type}")

        move = _norm(read_text(row, "Move_Type"))
        gate_ts = read_timestamp(row, "Truck_In_Time")
        if move and gate_ts is not None:
            event_type = "GATE_OUT" if ("OUT" in move or "DELIVER" in move) else "GATE_IN"
            trail.append({"eventType": event_type, "ts": iso_utc(gate_ts)})
            ledger.observed(f"event:{event_type}", iso_utc(gate_ts),
                            "Move_Type + Truck_In_Time", move,
                            f"{move} -> {event_type}")

    # De-duplicate: the same event can appear on more than one source row.
    seen: set = set()
    unique_trail = []
    for entry in sorted(trail, key=lambda e: e["ts"]):
        key = (entry["eventType"], entry["ts"])
        if key not in seen:
            seen.add(key)
            unique_trail.append(entry)

    if not unique_trail:
        ledger.warn("no recognisable lifecycle timestamps in the supplied rows")
        return _merge_degraded({
            "moduleId": m4.MODULE_ID, "container": container,
            "eventCount": 0, "clean": True, "findingCount": 0,
            "findings": [], "worstSeverity": None,
            "degraded": True,
            "decision_path": "engine=rule_engine | no_events_extracted",
        }, ledger)

    now_dt = read_timestamp({"now": now}, "now") if now else None
    result = m4.evaluate_trail(unique_trail, now=now_dt, container=container)
    response = result.as_dict()
    response["trailUsed"] = unique_trail
    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 9 -- J10  ->  M5 DISCHARGE AND BERTH STAY
# ==========================================================================

def reforecast_j10_vessel(row: Dict[str, Any],
                          moves_done: Optional[int] = None,
                          elapsed_h: Optional[float] = None) -> Dict[str, Any]:
    """
    Map one J10_Vessel_Call_Berthing row to M5 and project the berth stay.

    J10 gives the call, the terminal and -- usefully -- ``Cranes_Allocated``,
    which M5 needs and previously had to default. What J10 cannot give is live
    progress: ``moves_done`` and ``elapsed_h`` come from the TOS, so the caller
    passes them. Without them the response is a PLAN projection at the
    terminal's assumed crane rate, and it says so.
    """
    ledger = MappingLedger("J10_Vessel_Call_Berthing", "UC2-M5")

    via = read_text(row, "VIA_Visit", "VCN", "Voyage_No") or "UNKNOWN"
    ledger.observed("via_no", via, "VIA_Visit", via)

    terminal = read_text(row, "Terminal_Code") or "DEFAULT"
    if terminal.upper() not in m5.TERMINAL_CRANE_PRODUCTIVITY:
        ledger.warn(f"terminal {terminal!r} has no measured crane productivity; "
                    f"M5 will use its DEFAULT of "
                    f"{m5.TERMINAL_CRANE_PRODUCTIVITY['DEFAULT']} moves/crane-hour")
    ledger.observed("terminal", terminal, "Terminal_Code", terminal)

    export_boxes = read_int(row, "Export_Containers_On_EAL") or 0
    import_boxes = read_int(row, "Import_Containers_Manifested") or 0
    total = export_boxes + import_boxes
    if total > 0:
        moves_total = ledger.observed(
            "moves_total", total, "Export_Containers_On_EAL + "
            "Import_Containers_Manifested", f"{export_boxes}+{import_boxes}",
            "declared box counts; one move per box")
    else:
        moves_total = ledger.assumed(
            "moves_total", 1200,
            "neither Export_Containers_On_EAL nor Import_Containers_Manifested "
            "supplied; 1200 is the corpus median call size")

    cranes = read_float(row, "Cranes_Allocated", "Crane_Count_Allocated")
    if cranes is not None and cranes > 0:
        cranes = ledger.observed("cranes", cranes, "Cranes_Allocated", cranes)
    else:
        cranes = ledger.assumed("cranes", DEFAULT_CRANES,
                                "no Cranes_Allocated column")

    if moves_done is None:
        moves_done = ledger.assumed(
            "moves_done", 0,
            "live progress is a TOS field, not a J10 column; with 0 moves done "
            "the response is a plan projection at the assumed crane rate, not "
            "a re-forecast from observed progress")
    else:
        ledger.observed("moves_done", moves_done, "caller (TOS)", moves_done)

    if elapsed_h is None:
        eta = read_timestamp(row, "ETA", "Gate_Open_DateTime")
        elapsed_h = ledger.assumed(
            "elapsed_h", 0.0,
            "no live elapsed time supplied; the projection is from berthing, "
            "not from current progress")
    else:
        ledger.observed("elapsed_h", elapsed_h, "caller (TOS)", elapsed_h)

    planned = read_float(row, "Planned_Stay_H", "Planned_Berthing_Hours",
                         "Berth_Plan_Hours")
    if planned is not None:
        ledger.observed("planned_stay_h", planned, "Planned_Stay_H", planned,
                        "declared berth plan")
    else:
        # DO NOT infer the plan from ETA -> Declared_Sailing_DateTime.
        #
        # On the JNPA rows those two bracket the EXPORT RECEIVING window, not
        # the berth stay: NORTHERN PRACTISE has Gate_Open == ETA == 10 Jun
        # 02:01 and sailing 15 Jun 08:01, a 126 h span, while the same call's
        # box count projects a ~20 h stay. Treating 126 h as the plan produces
        # a -106 h variance and an "AHEAD" badge -- a number an operator would
        # act on and be wrong. J2 confirms the reading: it carries the same
        # span as Hours_Gate_In_To_Cutoff (138.6 h), a cutoff, not a stay.
        #
        # So the window is reported for context under its own name and the
        # status badge is left UNKNOWN until a real berth plan is supplied.
        eta = read_timestamp(row, "ETA")
        sailing = read_timestamp(row, "Declared_Sailing_DateTime")
        if eta is not None and sailing is not None and sailing > eta:
            window_h = (sailing - eta).total_seconds() / 3600.0
            ledger.warn(
                f"ETA to Declared_Sailing_DateTime spans {window_h:.1f} h, but "
                f"that is the export receiving/cutoff window, NOT a berth plan. "
                f"It is reported as declaredWindowHours and is deliberately not "
                f"used as planned_stay_h -- supply Planned_Stay_H for a variance "
                f"and a status badge.")
        else:
            ledger.warn("no Planned_Stay_H column; the response carries no "
                        "variance-vs-plan and no status badge")

    result = m5.reforecast(
        via_no=via, terminal=terminal, moves_total=int(moves_total),
        moves_done=int(max(0, min(moves_done, moves_total))),
        elapsed_h=float(max(0.0, elapsed_h)),
        planned_stay_h=planned, cranes=float(cranes),
    )
    response = result.as_dict()
    response["vesselName"] = read_text(row, "Vessel_Name") or ""
    response["imo"] = read_text(row, "IMO") or ""

    eta = read_timestamp(row, "ETA")
    sailing = read_timestamp(row, "Declared_Sailing_DateTime")
    if eta is not None:
        response["etaUtc"] = iso_utc(eta)
        response["projectedDepartureUtc"] = iso_utc(
            eta + timedelta(hours=result.projected_total_stay_h))
    if eta is not None and sailing is not None and sailing > eta:
        response["declaredSailingUtc"] = iso_utc(sailing)
        response["declaredWindowHours"] = round(
            (sailing - eta).total_seconds() / 3600.0, 2)
        response["declaredWindowNote"] = (
            "ETA to declared sailing. This is the export receiving/cutoff "
            "window, not the berth plan -- do not render it as planned stay.")

    if moves_total < 50:
        ledger.warn(
            f"moves_total is only {int(moves_total)}; a J10 row whose box count "
            f"comes from a scan list rather than the full manifest will "
            f"under-project the stay. Check Import_Containers_Manifested.")

    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 10 -- J3 / J4 / J8 / J10  ->  M6 LANE ASSIGNMENT
# ==========================================================================

def plan_lanes_from_rows(gate_rows: Sequence[Dict[str, Any]],
                         closed_lanes: Sequence[str] = (),
                         vessel_rows: Sequence[Dict[str, Any]] = (),
                         ) -> Dict[str, Any]:
    """
    Build M6's per-class demand from real gate rows and plan the lanes.

    Two sources are combined, and the response says which contributed what:
        gate_rows    J3 transactions or J4 buckets -- the total truck demand
                     and, via Move_Type, how it splits across classes
        vessel_rows  J10 calls -- Reefer_Count and Hazardous_Count, the two
                     classes that decide whether the plan can execute at all

    Reefer and hazardous volumes matter far beyond their size: L5 is the only
    powered lane and L6 the only hazmat lane, so a handful of boxes with the
    wrong lane closed turns the whole plan unservable.
    """
    ledger = MappingLedger("J3/J4 gate rows + J10 vessel rows", "UC2-M6")
    demand: Dict[str, float] = {c: 0.0 for c in m6.MOVEMENT_CLASSES}

    # A J4 bucket row states an HOURLY RATE (Trucks_In_Last_Hour). A J3 row is
    # ONE TRANSACTION. Summing them the same way is the mistake that makes a
    # saturated gate look idle, so the two are separated and the class SHARE is
    # taken from the transaction mix while the TOTAL comes from the rate rows.
    rate_rows = [r for r in gate_rows if read_float(r, "Trucks_In_Last_Hour")]
    txn_rows = [r for r in gate_rows if not read_float(r, "Trucks_In_Last_Hour")]

    def classify(row: Dict[str, Any]) -> Optional[str]:
        if _norm(read_text(row, "Container_Status")) in {"MTY", "EMPTY"}:
            return "EMPTY"
        move = _norm(read_text(row, "Move_Type"))
        if move and move in MOVE_TYPE_TO_CLASS:
            return MOVE_TYPE_TO_CLASS[move]
        if move:
            ledger.warn(f"Move_Type {move!r} is not in MOVE_TYPE_TO_CLASS; the "
                        f"row is left unassigned rather than guessed into a class")
        return None

    # Class shares from whichever rows carry a Move_Type at all.
    share_counts: Dict[str, int] = {}
    for row in gate_rows:
        cls = classify(row)
        if cls:
            share_counts[cls] = share_counts.get(cls, 0) + 1
    total_classified = sum(share_counts.values())

    total_per_hour = 0.0
    if rate_rows:
        # Each rate row already states trucks/hour for its gate-hour bucket.
        # Averaging across gates would understate a multi-gate terminal, so
        # they are summed and the row count is reported.
        total_per_hour = sum(read_float(r, "Trucks_In_Last_Hour") or 0.0
                             for r in rate_rows)
        ledger.observed(
            "total_demand_per_hour", round(total_per_hour, 2),
            "Trucks_In_Last_Hour", f"{len(rate_rows)} gate-hour bucket(s)",
            "summed observed hourly truck arrivals across the supplied buckets")
    elif txn_rows:
        # Transactions: count them per clock hour and take the busiest hour,
        # because a lane plan has to survive the peak, not the daily mean.
        by_hour: Dict[Any, int] = {}
        for row in txn_rows:
            hour = read_int(row, "Hour_Of_Day")
            if hour is None:
                ts = read_timestamp(row, "Truck_In_Time", "Timestamp")
                hour = ts.hour if ts else None
            key = (read_text(row, "Terminal_Code") or "", hour)
            by_hour[key] = by_hour.get(key, 0) + 1
        total_per_hour = float(max(by_hour.values())) if by_hour else 0.0
        ledger.observed(
            "total_demand_per_hour", round(total_per_hour, 2),
            "count of J3 transactions per gate-hour", dict(list(by_hour.items())[:6]),
            "busiest observed gate-hour; a lane plan must survive the peak, "
            "not the daily mean")
        ledger.warn(
            f"demand was counted from {len(txn_rows)} individual gate "
            f"transactions, which is a sample of the hour rather than a "
            f"measured hourly rate. Send J4 buckets with Trucks_In_Last_Hour "
            f"for a measured total.")

    if total_classified and total_per_hour > 0:
        for cls, count in share_counts.items():
            demand[cls] += total_per_hour * count / total_classified
        ledger.observed(
            "class_shares", {k: round(v / total_classified, 3)
                             for k, v in share_counts.items()},
            "Move_Type / Container_Status", share_counts,
            "share of the hourly total sent to each movement class")
    elif total_per_hour > 0:
        ledger.warn("no row carried a usable Move_Type; the hourly total could "
                    "not be split across movement classes and no lane plan can "
                    "be produced from it")

    classified = total_classified

    reefer_total = 0.0
    hazmat_total = 0.0
    for row in vessel_rows:
        reefer_total += read_float(row, "Reefer_Count") or 0.0
        hazmat_total += read_float(row, "Hazardous_Count") or 0.0

    if reefer_total or hazmat_total:
        # A call's boxes evacuate over roughly a day, so per-hour demand is
        # the count spread across 24 h. Named here, not hidden.
        demand["REEFER"] += reefer_total / 24.0
        demand["HAZARDOUS"] += hazmat_total / 24.0
        ledger.observed(
            "reefer_hazmat_demand",
            {"REEFER": round(reefer_total / 24.0, 3),
             "HAZARDOUS": round(hazmat_total / 24.0, 3)},
            "J10 Reefer_Count + Hazardous_Count",
            f"{reefer_total:g} reefer, {hazmat_total:g} hazmat",
            "counts spread over a 24 h evacuation window")

    if sum(demand.values()) <= 0:
        return _merge_degraded({
            "moduleId": m6.MODULE_ID,
            "error": "no demand could be derived from the supplied rows",
            "degraded": True,
            "decision_path": "adapter=no_demand_rows",
        }, ledger)

    demand = {k: round(v, 3) for k, v in demand.items() if v > 0}
    result = m6.assign_lanes(demand_per_hour=demand,
                             closed_lanes=list(closed_lanes))
    response = result.as_dict() if hasattr(result, "as_dict") else dict(result)
    response["demand_source"] = "caller_rows"
    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 11 -- J8 / J10  ->  M7 EMPTY POOL AND REEFER
# ==========================================================================

def allocate_reefers_from_rows(yard_rows: Sequence[Dict[str, Any]],
                               vessel_rows: Sequence[Dict[str, Any]] = (),
                               plugs_failed: Optional[int] = None,
                               cargo_rows: Sequence[Dict[str, Any]] = (),
                               ) -> Dict[str, Any]:
    """
    Allocate reefer plugs using the yard's REAL plug count.

    This is the mapping that most changes M7's answer. Its default plug count
    is 96 -- the Container Parking Plaza bank. J8 shows the terminals actually
    run 120 (NSFT), 150 (NSICT), 180 (BMCT) and 40 (the CFS). Summing the real
    banks instead of assuming 96 changes the shortfall, and therefore the
    countdown to the first box at risk.

    Plugs already in use are subtracted: a plug under a box is not available
    for the boxes about to land, and treating it as available would understate
    the exposure.
    """
    ledger = MappingLedger("J8_Yard_Empty_Pendency + J10", "UC2-M7")

    plugs_total = 0.0
    plugs_used = 0.0
    for row in yard_rows:
        total = read_float(row, "Reefer_Plugs_Total")
        used = read_float(row, "Reefer_Plugs_Used")
        if total:
            plugs_total += total
        if used:
            plugs_used += used

    if plugs_total > 0:
        ledger.observed("plugs_total", int(plugs_total), "Reefer_Plugs_Total",
                        plugs_total,
                        f"summed across {len(yard_rows)} yard block(s); "
                        f"M7's own default is {m7.CPP_REEFER_PLUGS}")
    else:
        plugs_total = ledger.assumed(
            "plugs_total", m7.CPP_REEFER_PLUGS,
            "no Reefer_Plugs_Total column; falling back to the Container "
            "Parking Plaza bank")

    if plugs_used > 0:
        ledger.observed("plugs_in_use", int(plugs_used), "Reefer_Plugs_Used",
                        plugs_used,
                        "occupied plugs are unavailable to arriving boxes and "
                        "are counted as failed for allocation purposes")

    reefers = 0.0
    for row in vessel_rows:
        reefers += read_float(row, "Reefer_Count") or 0.0
    if reefers > 0:
        ledger.observed("reefers_arriving", int(reefers), "Reefer_Count",
                        reefers, f"summed across {len(vessel_rows)} vessel call(s)")
    else:
        reefers = ledger.assumed("reefers_arriving", 200,
                                 "no Reefer_Count on any supplied vessel row")

    if plugs_failed is None:
        plugs_failed = ledger.assumed(
            "plugs_failed", 0,
            "plug faults are an operations input, not a sheet column; "
            "assuming every plug in the bank is live")

    # Occupied plugs are unavailable to the arriving parcel, exactly like a
    # failed one. Both are folded into plugs_failed, and the ledger says so.
    unavailable = int(min(plugs_total, plugs_used + plugs_failed))

    mix = resolve_sensitivity_mix(cargo_rows, ledger) if cargo_rows else None

    result = m7.allocate_reefer_plugs(
        reefers_arriving=int(reefers),
        sensitivity_mix=mix,
        plugs_failed=unavailable,
        plugs_total=int(plugs_total),
        scenario_id="WEBAPP",
        title="Reefer allocation from live yard and vessel rows",
    )
    response = result.as_dict()
    response["plugsInUse"] = int(plugs_used)
    response["plugsFailedReported"] = int(plugs_failed)
    response["plugsUnavailableTotal"] = unavailable
    return _merge_degraded(response, ledger)


def balance_empties_from_rows(yard_rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Empty-pool balance from J8's real per-terminal empty counts.

    J8 carries ``Empty_Boxes`` on hand and ``Empties_Out_Last_24h`` -- an
    observed daily burn rate. That is a far better demand signal than M7's
    export-laden proxy, so it is used when present and the response says which
    route produced the numbers.
    """
    ledger = MappingLedger("J8_Yard_Empty_Pendency", "UC2-M7")

    daily_demand: Dict[str, float] = {}
    on_hand: Dict[str, float] = {}
    for row in yard_rows:
        terminal = read_text(row, "Terminal_Code", "Yard_Block")
        if not terminal:
            continue
        out_24h = read_float(row, "Empties_Out_Last_24h")
        empties = read_float(row, "Empty_Boxes")
        if out_24h is not None and out_24h > 0:
            daily_demand[terminal] = daily_demand.get(terminal, 0.0) + out_24h
        if empties is not None:
            on_hand[terminal] = on_hand.get(terminal, 0.0) + empties

    if daily_demand:
        ledger.observed("daily_demand_by_terminal",
                        {k: round(v, 1) for k, v in daily_demand.items()},
                        "Empties_Out_Last_24h", daily_demand,
                        "observed 24 h empty outflow IS the daily demand; "
                        "this replaces M7's export-laden proxy")
    else:
        ledger.warn("no Empties_Out_Last_24h on any row; M7 will estimate "
                    "demand from export-laden volumes in the corpus")

    snapshot = None
    if on_hand:
        # Build the supply side from the SAME J8 rows the demand came from.
        #
        # Without this M7 reads supply from its own corpus inventory while the
        # demand comes from JNPA's yard snapshot, and the two do not describe
        # the same day or the same terminals: the mixed calculation gave GTI
        # 1662 days of cover and NSFT zero, neither of which appears in either
        # dataset. One source per calculation, or the answer is fiction.
        by_terminal: Dict[str, Dict[str, int]] = {}
        for row in yard_rows:
            terminal = read_text(row, "Terminal_Code", "Yard_Block")
            if not terminal:
                continue
            bucket = by_terminal.setdefault(
                terminal, {"total": 0, "empty": 0, "laden": 0, "reefer": 0,
                           "import": 0, "export": 0})
            empty = int(read_float(row, "Empty_Boxes") or 0)
            imports = int(read_float(row, "Import_Boxes") or 0)
            exports = int(read_float(row, "Export_Boxes") or 0)
            reefers = int(read_float(row, "Reefer_Plugs_Used") or 0)
            bucket["empty"] += empty
            bucket["import"] += imports
            bucket["export"] += exports
            bucket["laden"] += imports + exports
            bucket["reefer"] += reefers
            bucket["total"] += empty + imports + exports

        snapshot = m7.PoolSnapshot(
            source="WEBAPP_J8",
            total_containers=sum(b["total"] for b in by_terminal.values()),
            empties=sum(b["empty"] for b in by_terminal.values()),
            reefers=sum(b["reefer"] for b in by_terminal.values()),
            by_terminal=by_terminal,
            by_line={},
            by_size={},
            provenance={
                "source": "J8_Yard_Empty_Pendency rows supplied by the caller",
                "rows": len(yard_rows),
                "terminals": sorted(by_terminal),
                "note": ("Supply and demand both come from these rows, so the "
                         "days-of-cover figure describes one consistent "
                         "snapshot rather than mixing two datasets."),
            },
        )
        ledger.observed(
            "empties_on_hand", {k: int(v) for k, v in on_hand.items()},
            "Empty_Boxes", on_hand,
            "supply side of the balance, taken from the same rows as demand")

    result = m7.balance_empty_pool(
        daily_demand_by_terminal=daily_demand or None, snapshot=snapshot)
    response = result.as_dict() if hasattr(result, "as_dict") else dict(result)
    response["supplySource"] = "WEBAPP_J8" if snapshot else "CORPUS"
    if snapshot is None:
        ledger.warn("no Empty_Boxes column on any row; the supply side comes "
                    "from M7's corpus inventory while demand came from your "
                    "rows -- send Empty_Boxes so both sides agree")
    return _merge_degraded(response, ledger)


# ==========================================================================
# SECTION 12 -- MAPPING CATALOGUE (served to the frontend)
# ==========================================================================

def mapping_catalogue() -> Dict[str, Any]:
    """
    Every code table this adapter validates against.

    The frontend renders its dropdowns from this so a value the user can pick
    is always a value the adapter can map.
    """
    return {
        "adapter_version": MODULE_VERSION,
        "streams": {i: name for i, name in enumerate(m1.STREAMS)},
        "lines": {i: name for i, name in enumerate(m1.LINES)},
        "line_codes": LINE_CODE_TO_IDX,
        "delivery_modes": DELIVERY_MODE_MEANING,
        "sidings": SIDING_TO_IDX,
        "ctos": {i: name for i, name in enumerate(m2.CTOS)},
        "cto_codes": CTO_TO_IDX,
        "movement_classes": list(m6.MOVEMENT_CLASSES),
        "move_types": MOVE_TYPE_TO_CLASS,
        "sensitivity_hold_hours": m7.HOLD_HOURS_BY_SENSITIVITY,
        "sensitivity_priority": list(m7.SENSITIVITY_PRIORITY),
        "terminal_crane_productivity": m5.TERMINAL_CRANE_PRODUCTIVITY,
        "j9_flag_rules": FLAG_FINDINGS,
        "reefer_text_markers": list(REEFER_TEXT_MARKERS),
        "customs_held_states": list(CUSTOMS_HELD_STATES),
        "defaults": {
            "arrival_cadence_h": DEFAULT_ARRIVAL_CADENCE_H,
            "facility_load": DEFAULT_FACILITY_LOAD,
            "terminal_count": DEFAULT_TERMINAL_COUNT,
            "cranes": DEFAULT_CRANES,
            "zscore_outlier_threshold": ZSCORE_OUTLIER_THRESHOLD,
        },
        "note": (
            "Every default above sets degraded=true and is named in "
            "mapping.assumptions[] whenever it is used. Nothing is silently "
            "substituted."
        ),
    }


# ==========================================================================
# SECTION 13 -- HTTP ROUTER
# ==========================================================================

try:
    from pydantic import BaseModel, Field

    class TrailRequest(BaseModel):
        """Every lifecycle row the caller holds for one container."""

        rows: List[JnpaRow] = Field(..., min_length=1)
        container: str = ""
        now: Optional[str] = None

    class LanePlanRequest(BaseModel):
        """Gate rows for demand, vessel rows for reefer/hazmat, closures."""

        gate_rows: List[JnpaRow] = Field(default_factory=list)
        vessel_rows: List[JnpaRow] = Field(default_factory=list)
        closed_lanes: List[str] = Field(default_factory=list)

    class ReeferRequest(BaseModel):
        """Yard rows for the plug bank, vessel rows for the arriving parcel."""

        yard_rows: List[JnpaRow] = Field(default_factory=list)
        vessel_rows: List[JnpaRow] = Field(default_factory=list)
        cargo_rows: List[JnpaRow] = Field(default_factory=list)
        # Bounded by the bank being described, not by the Container Parking
        # Plaza's 96 -- JNPA yards run up to 180 plugs.
        plugs_failed: Optional[int] = Field(None, ge=0, le=1000)

    class VesselRequest(BaseModel):
        """One J10 call, plus live TOS progress when the caller has it."""

        row: JnpaRow
        moves_done: Optional[int] = Field(None, ge=0, le=100000)
        elapsed_h: Optional[float] = Field(None, ge=0, le=500)

except ImportError:  # pragma: no cover - CLI path needs no FastAPI/pydantic
    TrailRequest = LanePlanRequest = ReeferRequest = VesselRequest = None  # type: ignore


def build_router():  # pragma: no cover - exercised by the service
    """FastAPI router exposing the adapter under /uc2/webapp."""
    from fastapi import APIRouter, Body, HTTPException, Query

    router = APIRouter(prefix=ROUTER_PREFIX, tags=["UC2-ADAPTER (web app ingest)"])

    def _guard(fn, *args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

    @router.get("/mappings", summary="Every code table the adapter validates against")
    def mappings() -> Dict[str, Any]:
        return mapping_catalogue()

    @router.post("/m1/dwell", summary="J1/J2/J6 row -> container dwell forecast")
    def m1_dwell(row: JnpaRow = Body(...)) -> Dict[str, Any]:
        return _guard(predict_j1_dwell, row)

    @router.post("/m1/dwell-batch",
                 summary="Many lifecycle rows -> forecasts with MEASURED cadence")
    def m1_dwell_batch(rows: List[JnpaRow] = Body(...)) -> Dict[str, Any]:
        # Prefer this over /m1/dwell in a loop: arrival cadence is only
        # measurable across rows, so the batch path degrades far less.
        return _guard(predict_dwell_batch, rows)

    @router.post("/m2/rake", summary="J7 row -> rake turnaround forecast")
    def m2_rake(row: JnpaRow = Body(...),
                engine: str = Query("handling", pattern="^(handling|learned)$")
                ) -> Dict[str, Any]:
        return _guard(predict_j7_rake, row, engine)

    @router.post("/m3/queue", summary="J3/J4 row -> next-hour gate queue")
    def m3_queue(row: JnpaRow = Body(...)) -> Dict[str, Any]:
        return _guard(predict_j4_queue, row)

    @router.post("/m3/queue-curve", summary="J4 row -> multi-hour queue curve")
    def m3_curve(row: JnpaRow = Body(...),
                 hours: int = Query(12, ge=1, le=72)) -> Dict[str, Any]:
        return _guard(forecast_j4_curve, row, hours)

    @router.post("/m4/event", summary="J9 row -> anomaly findings (flag passthrough)")
    def m4_event(row: JnpaRow = Body(...)) -> Dict[str, Any]:
        return _guard(evaluate_j9_event, row)

    @router.post("/m4/trail", summary="Lifecycle rows -> M4 R1-R6 trail rules")
    def m4_trail(req: TrailRequest) -> Dict[str, Any]:
        return _guard(evaluate_container_trail, req.rows, req.container, req.now)

    @router.post("/m5/vessel", summary="J10 row -> berth stay projection")
    def m5_vessel(req: VesselRequest) -> Dict[str, Any]:
        return _guard(reforecast_j10_vessel, req.row, req.moves_done, req.elapsed_h)

    @router.post("/m6/lanes", summary="Gate + vessel rows -> lane assignment plan")
    def m6_lanes(req: LanePlanRequest) -> Dict[str, Any]:
        return _guard(plan_lanes_from_rows, req.gate_rows, req.closed_lanes,
                      req.vessel_rows)

    @router.post("/m7/reefer", summary="J8 + J10 rows -> reefer plug allocation")
    def m7_reefer(req: ReeferRequest) -> Dict[str, Any]:
        return _guard(allocate_reefers_from_rows, req.yard_rows, req.vessel_rows,
                      req.plugs_failed, req.cargo_rows)

    @router.post("/m7/empties", summary="J8 rows -> empty pool balance")
    def m7_empties(yard_rows: List[JnpaRow] = Body(...)) -> Dict[str, Any]:
        return _guard(balance_empties_from_rows, yard_rows)

    @router.get("/health", summary="Adapter health")
    def health() -> Dict[str, Any]:
        checks = selftest()
        return {
            "moduleId": MODULE_ID, "version": MODULE_VERSION,
            "ok": all(c["passed"] for c in checks), "checks": checks,
        }

    return router


# ==========================================================================
# SECTION 14 -- SELF-TEST
# ==========================================================================

def selftest() -> List[Dict[str, Any]]:
    """Checks that the translations this module exists to perform are right."""
    checks: List[Dict[str, Any]] = []

    def check(name: str, condition: bool, detail: str = "") -> None:
        checks.append({"check": name, "passed": bool(condition), "detail": detail})

    # The three translations a frontend would most plausibly get wrong.
    ledger = MappingLedger("test", "UC2-M2")
    check("siding T2 -> 1", SIDING_TO_IDX["T2"] == 1)
    check("CTO-2 -> index 1 (sheet 1-based, model 0-based)",
          CTO_TO_IDX["CTO-2"] == 1,
          f"CTO-2 is {m2.CTOS[1]}, not {m2.CTOS[2]}")
    check("CTO-1 -> CONCOR", m2.CTOS[CTO_TO_IDX["CTO-1"]] == "CONCOR")

    # An unrecognised line must not silently become MSC.
    idx = resolve_line_idx({"Shipping_Line_Code": "ZZZZ"}, MappingLedger("t", "t"))
    check("unknown line code -> OTHER, not MSC",
          m1.LINES[idx] == "OTHER", f"got {m1.LINES[idx]}")

    # Container prefix fallback must work when the line code is missing.
    idx = resolve_line_idx({"Container_No": "MSKU2256091"}, MappingLedger("t", "t"))
    check("ISO owner prefix MSKU -> MAERSK", m1.LINES[idx] == "MAERSK",
          f"got {m1.LINES[idx]}")

    # Empty boxes must classify as EMPTY_RETURN whatever else the row says.
    idx = resolve_stream_idx({"Container_Status": "MTY", "Delivery_Mode": "G"},
                             MappingLedger("t", "t"))
    check("MTY -> EMPTY_RETURN", m1.STREAMS[idx] == "EMPTY_RETURN")

    # PENDING is not a hold; HELD is.
    led = MappingLedger("t", "t")
    check("Customs_Status PENDING is not a hold",
          resolve_customs_flag({"Customs_Status": "PENDING"}, led) == 0)
    check("Customs_Status HELD is a hold",
          resolve_customs_flag({"Customs_Status": "HELD"}, led) == 1)
    check("Selected_Scan Yes is a hold",
          resolve_customs_flag({"Selected_Scan": "Yes"}, led) == 1)

    # Percent must become a fraction, and a bad column must not silently zero.
    led = MappingLedger("t", "t")
    check("yard utilisation 74% -> 0.74",
          abs(resolve_facility_load({"Terminal_Yard_Utilization_Pct": 74}, led)
              - 0.74) < 1e-9)
    led = MappingLedger("t", "t")
    check("missing yard utilisation is assumed, not zeroed",
          resolve_facility_load({}, led) == DEFAULT_FACILITY_LOAD and led.degraded)

    # Timestamp shapes from the three upstream systems.
    check("'2026-06-07 08:27' parses",
          read_timestamp({"t": "2026-06-07 08:27"}, "t") is not None)
    check("'07/06/2026 08:27' parses",
          read_timestamp({"t": "07/06/2026 08:27"}, "t") is not None)
    check("'07062026:08:27' (EDI) parses",
          read_timestamp({"t": "07062026:08:27"}, "t") is not None)

    # Booleans in every shape the sheets use.
    check("'Yes' is true", read_bool({"t": "Yes"}, "t") is True)
    check("'No' is false", read_bool({"t": "No"}, "t") is False)
    check("blank is None, not false", read_bool({"t": ""}, "t") is None)

    # An assumed input must set degraded.
    led = MappingLedger("t", "t")
    led.assumed("x", 1, "because")
    check("assumption sets degraded", led.degraded)
    led2 = MappingLedger("t", "t")
    led2.observed("x", 1, "col")
    check("observed input does not set degraded", not led2.degraded)

    # End-to-end on a real sample row.
    try:
        out = predict_j1_dwell({
            "Container_No": "DPWU9011100", "Shipping_Line_Code": "CHZ",
            "Terminal_Code": "NSICT", "Delivery_Mode": "G",
            "Customs_Status": "PENDING", "Nature_Of_Cargo": "GENERAL",
            "Terminal_Yard_Utilization_Pct": 74,
            "Arrival_DateTime": "2026-06-07 08:27", "Container_Status": "FCL",
            "IGM_No": "1194313",
        })
        check("J1 row predicts a positive dwell",
              out.get("p50Hours", 0) > 0, f"p50={out.get('p50Hours')}")
        check("J1 response carries a departure clock time",
              out.get("predictedDepartureUtc") is not None)
        check("J1 response carries the mapping ledger",
              "mapping" in out and out["mapping"]["derived"])
    except Exception as exc:  # noqa: BLE001
        check("J1 end-to-end", False, repr(exc))

    try:
        out = predict_j7_rake({
            "Rake_ID": "RK-AGM-0654", "Siding": "T2", "CTO_Index": "CTO-2",
            "Terminal_Code": "NSICT", "Direction": "Inbound",
            "Wagon_Count": 45, "Container_Count": 90, "Arrival_Hour": 14,
            "Destination_Terminal": "NSICT",
            "Arrival_Timestamp": "2026-06-11 14:00",
        })
        check("J7 row predicts a positive TAT",
              out.get("tatHours", 0) > 0, f"tat={out.get('tatHours')}")
        check("J7 maps siding T2 to index 1",
              any(d["model_input"] == "siding" and d["value"] == 1
                  for d in out["mapping"]["derived"]))
    except Exception as exc:  # noqa: BLE001
        check("J7 end-to-end", False, repr(exc))

    try:
        out = predict_j4_queue({
            "Gate_ID": "NSICT-G1", "Terminal_Code": "NSICT",
            "Timestamp": "2026-06-12 04:45", "Queue_Lag1": 9, "Queue_Lag2": 7,
            "Hour_Of_Day": 4, "UC3_Truck_Inflow_Per_Hr": 24, "Lanes_Open": 3,
        })
        check("J4 row predicts a non-negative queue",
              out.get("queueVehicles", -1) >= 0)
        check("J4 row with full lags is NOT degraded",
              out.get("degraded") is False,
              f"assumptions={out['mapping']['assumptions']}")
    except Exception as exc:  # noqa: BLE001
        check("J4 end-to-end", False, repr(exc))

    # Batch cadence. The shipped J1 sample cannot exercise this -- it carries
    # two arrival timestamps at two different terminals -- so the path is
    # proven here on rows that can.
    try:
        batch = predict_dwell_batch([
            {"Container_No": "AAAU1000001", "Terminal_Code": "NSICT",
             "Arrival_DateTime": "2026-06-07 08:00", "Delivery_Mode": "G",
             "Terminal_Yard_Utilization_Pct": 74, "Container_Status": "FCL",
             "Customs_Status": "PENDING", "Shipping_Line_Code": "MSC",
             "Nature_Of_Cargo": "GENERAL"},
            {"Container_No": "AAAU1000002", "Terminal_Code": "NSICT",
             "Arrival_DateTime": "2026-06-07 11:30", "Delivery_Mode": "G",
             "Terminal_Yard_Utilization_Pct": 74, "Container_Status": "FCL",
             "Customs_Status": "PENDING", "Shipping_Line_Code": "MSC",
             "Nature_Of_Cargo": "GENERAL"},
            {"Container_No": "AAAU1000003", "Terminal_Code": "NSICT",
             "Arrival_DateTime": "2026-06-07 13:00", "Delivery_Mode": "G",
             "Terminal_Yard_Utilization_Pct": 74, "Container_Status": "FCL",
             "Customs_Status": "PENDING", "Shipping_Line_Code": "MSC",
             "Nature_Of_Cargo": "GENERAL"},
        ])
        check("batch measures cadence for all but the first row",
              batch["cadence"]["measured_rows"] == 2
              and batch["cadence"]["assumed_rows"] == 1,
              f"measured={batch['cadence']['measured_rows']} "
              f"assumed={batch['cadence']['assumed_rows']}")
        second = batch["results"][1]
        cadence = [d for d in second["mapping"]["derived"]
                   if d["model_input"] == "arrival_cadence_h"][0]
        check("measured cadence is the real 3.5 h gap, not the 6.0 h default",
              abs(cadence["value"] - 3.5) < 1e-6 and cadence["observed"],
              f"got {cadence['value']} observed={cadence['observed']}")
        check("a row with measured cadence is not degraded",
              second["degraded"] is False,
              f"assumptions={second['mapping']['assumptions']}")
        check("the first row at a terminal is honestly marked degraded",
              batch["results"][0]["degraded"] is True)
    except Exception as exc:  # noqa: BLE001
        check("batch cadence", False, repr(exc))

    # M6 must not read a J3 transaction as an hourly rate.
    try:
        j4_plan = plan_lanes_from_rows([
            {"Gate_ID": "NSICT-G1", "Terminal_Code": "NSICT", "Hour_Of_Day": 4,
             "Trucks_In_Last_Hour": 31, "Move_Type": "Deliver Import",
             "Container_Status": "FCL"},
        ])
        check("a J4 bucket's Trucks_In_Last_Hour becomes the hourly demand",
              abs(j4_plan.get("totalDemandPerHour", 0) - 31.0) < 0.01,
              f"got {j4_plan.get('totalDemandPerHour')}")
    except Exception as exc:  # noqa: BLE001
        check("M6 rate-row demand", False, repr(exc))

    # Regression: M4 must not silently drop a Z-suffixed or space-separated
    # timestamp. It used to, and normalise_trail() drops unparseable rows
    # quietly -- so a real broken trail came back "clean, 0 events", which is
    # the one direction an anomaly detector must never fail in.
    try:
        rows = [
            {"Container_No": "ZZZU1000001", "Terminal_Code": "NSICT",
             "Arrival_DateTime": "2026-06-07 08:27"},          # sheet format
        ]
        trail_out = evaluate_container_trail(rows, container="ZZZU1000001",
                                             now="2026-06-20T08:00:00Z")
        check("M4 trail parses the sheet's 'YYYY-MM-DD HH:MM' timestamps",
              trail_out.get("eventCount", 0) == 1,
              f"eventCount={trail_out.get('eventCount')}")
        check("a 13-day-old gate-in with no gate-out raises R1",
              any(f["ruleId"] == "R1" for f in trail_out.get("findings", [])),
              f"findings={[f['ruleId'] for f in trail_out.get('findings', [])]}")
    except Exception as exc:  # noqa: BLE001
        check("M4 trail timestamp parsing", False, repr(exc))

    # M7's plug ceiling must accept a real JNPA yard bank, not just 96.
    try:
        alloc = allocate_reefers_from_rows(
            yard_rows=[{"Terminal_Code": "BMCT", "Reefer_Plugs_Total": 180,
                        "Reefer_Plugs_Used": 141}],
            vessel_rows=[{"Reefer_Count": 44}], plugs_failed=100)
        check("M7 accepts a 180-plug bank with 100 faults (was capped at 96)",
              alloc.get("plugsTotal") == 180,
              f"plugsTotal={alloc.get('plugsTotal')}")
    except Exception as exc:  # noqa: BLE001
        check("M7 large plug bank", False, repr(exc))

    return checks


# ==========================================================================
# SECTION 15 -- CLI
# ==========================================================================

SAMPLE_WORKBOOK = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "data", "input", "uc2", "Cargo_Training_Input_Sample.xlsx")


def load_sample_sheets(path: str = SAMPLE_WORKBOOK) -> Dict[str, List[Dict[str, Any]]]:
    """Read the JNPA sample workbook into {sheet: [row dicts]}."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read the sample workbook") from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets: Dict[str, List[Dict[str, Any]]] = {}
    for name in workbook.sheetnames:
        if name in {"README", "Feature_Provenance"}:
            continue
        worksheet = workbook[name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        sheets[name] = [
            {h: v for h, v in zip(header, r) if h is not None}
            for r in rows[1:] if any(v is not None for v in r)
        ]
    return sheets


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=MODULE_NAME)
    parser.add_argument("--sheet", help="only this sheet (e.g. J1, J7)")
    parser.add_argument("--json", action="store_true", help="machine-readable")
    parser.add_argument("--selftest", action="store_true")
    parser.add_argument("--workbook", default=SAMPLE_WORKBOOK)
    args = parser.parse_args(argv)

    if args.selftest:
        checks = selftest()
        failed = [c for c in checks if not c["passed"]]
        if args.json:
            print(json.dumps({"checks": checks, "failed": len(failed)}, indent=2))
        else:
            for check in checks:
                mark = "PASS" if check["passed"] else "FAIL"
                detail = f"  [{check['detail']}]" if check["detail"] else ""
                print(f"  {mark}  {check['check']}{detail}")
            print(f"\n{len(checks) - len(failed)}/{len(checks)} checks passed")
        return 1 if failed else 0

    sheets = load_sample_sheets(args.workbook)
    results: Dict[str, Any] = {}

    def wanted(name: str) -> bool:
        return not args.sheet or args.sheet.upper() in name.upper()

    if wanted("J1_Import_Lifecycle"):
        results["M1_from_J1"] = predict_dwell_batch(
            sheets.get("J1_Import_Lifecycle", []))["results"]
    if wanted("J7_Rail_Rake"):
        results["M2_from_J7"] = [predict_j7_rake(r)
                                 for r in sheets.get("J7_Rail_Rake", [])]
    if wanted("J4_Gate_Queue_Forecast"):
        results["M3_from_J4"] = [predict_j4_queue(r)
                                 for r in sheets.get("J4_Gate_Queue_Forecast", [])]
    if wanted("J9_Event_Anomaly"):
        results["M4_from_J9"] = [evaluate_j9_event(r)
                                 for r in sheets.get("J9_Event_Anomaly", [])]
    if wanted("J10_Vessel_Call_Berthing"):
        results["M5_from_J10"] = [reforecast_j10_vessel(r)
                                  for r in sheets.get("J10_Vessel_Call_Berthing", [])]
    if wanted("J3_Gate_Transaction"):
        # J4 buckets carry Trucks_In_Last_Hour -- a measured hourly rate.
        # J3 transactions carry Move_Type -- the class mix. Both are needed:
        # J3 alone can only tell you the busiest sampled hour had one truck.
        results["M6_from_J3_J4_J10"] = plan_lanes_from_rows(
            list(sheets.get("J4_Gate_Queue_Forecast", []))
            + list(sheets.get("J3_Gate_Transaction", [])),
            closed_lanes=[],
            vessel_rows=sheets.get("J10_Vessel_Call_Berthing", []))
        results["M6_from_J3_J4_J10_hazmat_lane_closed"] = plan_lanes_from_rows(
            list(sheets.get("J4_Gate_Queue_Forecast", []))
            + list(sheets.get("J3_Gate_Transaction", [])),
            closed_lanes=["L6"],
            vessel_rows=sheets.get("J10_Vessel_Call_Berthing", []))
    if wanted("J8_Yard_Empty_Pendency"):
        results["M7_reefer_from_J8_J10"] = allocate_reefers_from_rows(
            sheets.get("J8_Yard_Empty_Pendency", []),
            sheets.get("J10_Vessel_Call_Berthing", []))
        results["M7_empties_from_J8"] = balance_empties_from_rows(
            sheets.get("J8_Yard_Empty_Pendency", []))

    if args.json:
        print(json.dumps(results, indent=2, default=str))
        return 0

    print(f"\n{MODULE_ID} {MODULE_VERSION} -- predictions from {os.path.basename(args.workbook)}\n")
    for key, value in results.items():
        print(f"  {key}: "
              f"{len(value) if isinstance(value, list) else 1} result(s)")
    print("\nRun with --json for the full payloads.")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
